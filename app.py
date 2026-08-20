#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OnFire — Plataforma de generación y control de boletos con QR
Backend: Flask + SQLite (archivo local, fácil de migrar a otra BD después).
Todos los datos de venta se sincronizan automáticamente a data/boletos.xlsx.
"""
import os, re, json, time, base64, shutil, sqlite3, secrets, hashlib, threading
from datetime import datetime, timedelta
from io import BytesIO
try:
    from zoneinfo import ZoneInfo
    EVENT_TZ = ZoneInfo(os.environ.get("EVENT_TZ", "America/Mexico_City"))
except Exception:
    EVENT_TZ = None   # sin base de zonas → cae a la hora local del servidor

from flask import Flask, request, jsonify, send_from_directory, send_file, g, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE    = os.path.dirname(os.path.abspath(__file__))
# DATA_DIR permite apuntar a un Volumen persistente (Railway u otro host).
# Si no se define, usa la carpeta local ./data (desarrollo).
DATA    = os.environ.get("DATA_DIR") or os.path.join(BASE, "data")
BACKUPS = os.path.join(DATA, "backups")
PUBLIC  = os.path.join(BASE, "public")
DB_PATH = os.path.join(DATA, "onfire.db")
XLSX    = os.path.join(DATA, "boletos.xlsx")

os.makedirs(BACKUPS, exist_ok=True)

app = Flask(__name__, static_folder=None)
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024  # flyer máx 8 MB

_write_lock = threading.Lock()

@app.after_request
def revalidate_assets(resp):
    # el navegador revalida HTML/JS/CSS en cada carga → nunca sirve una versión vieja
    ct = resp.headers.get("Content-Type", "")
    if any(t in ct for t in ("text/html", "javascript", "text/css")):
        resp.headers["Cache-Control"] = "no-cache"
    return resp

# ---------------------------------------------------------------- utilidades

def now_dt():
    return datetime.now(EVENT_TZ)   # hora del evento (México por defecto), no UTC del servidor

def now_iso():
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")

def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000)
    return f"{salt}${digest.hex()}"

def check_password(password, stored):
    try:
        salt, _ = stored.split("$", 1)
    except ValueError:
        return False
    return secrets.compare_digest(hash_password(password, salt), stored)

def money(cents):
    return float(cents or 0) / 100.0

# ------------------------------------------------------- contenido del QR
# El QR lleva un token aleatorio de 96 bits imposible de adivinar. El escáner lo
# valida contra la base en tiempo real; nadie puede fabricar un QR válido.
# Los datos legibles (nombre, tipo, folio) van IMPRESOS en el boleto, no en el QR.

def folio_from_scan(raw):
    """Del texto escaneado saca el identificador para buscar el boleto:
    el token tal cual, o el folio si se escribió/escaneó 'Folio HF-0001'."""
    raw = (raw or "").strip()
    m = re.search(r"[Ff]olio\s+(\S+)", raw)
    if m:
        return m.group(1)
    return raw

# ---------------------------------------------------------------- base de datos
# Funciona con SQLite (local, por defecto) o PostgreSQL (si existe DATABASE_URL,
# como en Railway). El resto del código usa la MISMA interfaz: db.execute("... ?",
# params).fetchone()/.fetchall(), db.commit(). El wrapper de Postgres traduce los
# marcadores ? → %s y entrega filas accesibles por nombre, igual que sqlite3.Row.

DATABASE_URL = os.environ.get("DATABASE_URL")
IS_PG = bool(DATABASE_URL)

if IS_PG:
    import psycopg
    from decimal import Decimal
    from psycopg.rows import dict_row
    IntegrityError = psycopg.IntegrityError
    LIKE = "ILIKE"   # búsqueda sin distinguir mayúsculas, como se comporta SQLite

    def _plain(row):
        """Convierte Decimal → int/float para que jsonify y el resto del código
        reciban los mismos tipos que con SQLite."""
        if row is None:
            return None
        out = {}
        for k, v in row.items():
            if isinstance(v, Decimal):
                v = int(v) if v == v.to_integral_value() else float(v)
            out[k] = v
        return out

    class _PGCursor:
        def __init__(self, cur, conn):
            self._cur, self._conn = cur, conn
        def fetchone(self):
            return _plain(self._cur.fetchone())
        def fetchall(self):
            return [_plain(r) for r in self._cur.fetchall()]
        @property
        def rowcount(self):
            return self._cur.rowcount
        @property
        def lastrowid(self):
            with self._conn.cursor() as c:
                c.execute("SELECT lastval()")
                return c.fetchone()[0]

    class PGConn:
        """Imita la interfaz de una conexión sqlite3 sobre psycopg."""
        def __init__(self, conn):
            self._conn = conn
        def execute(self, sql, params=()):
            cur = self._conn.cursor(row_factory=dict_row)
            # Sin parámetros se ejecuta TAL CUAL. Si se manda una tupla vacía, psycopg
            # igual busca marcadores y un '%' literal —el de un LIKE 'INV-%'— lo lee
            # como marcador roto y tira el arranque entero. En SQLite eso no pasa, así
            # que el error no se ve hasta que ya está en producción.
            if params:
                cur.execute(sql.replace("?", "%s"), params)
            else:
                cur.execute(sql)
            return _PGCursor(cur, self._conn)
        def executescript(self, script):
            # Se quitan los comentarios ANTES de partir por ";". Sin esto, un simple
            # punto y coma dentro de un comentario —"-- su comisión; NULL = ..."—
            # corta la sentencia a la mitad y el arranque revienta con "syntax error
            # at end of input", que no dice nada de dónde está el problema real.
            limpio = "\n".join(re.sub(r"--.*$", "", ln) for ln in script.splitlines())
            with self._conn.cursor() as cur:
                for stmt in limpio.split(";"):
                    if stmt.strip():
                        cur.execute(stmt)
        def commit(self):
            self._conn.commit()
        def rollback(self):
            self._conn.rollback()
        def close(self):
            self._conn.close()

    def db_connect():
        return PGConn(psycopg.connect(DATABASE_URL))
else:
    IntegrityError = sqlite3.IntegrityError
    LIKE = "LIKE"

    def db_connect():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

def get_db():
    if "db" not in g:
        g.db = db_connect()
    return g.db

@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

SCHEMA = """
CREATE TABLE IF NOT EXISTS settings (
  key TEXT PRIMARY KEY, value TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS admins (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT NOT NULL UNIQUE,
  pass_hash TEXT NOT NULL,
  role TEXT NOT NULL DEFAULT 'admin',   -- admin | colider (el colíder ve solo su grupo)
  created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS sellers (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  code TEXT,                       -- NULL cuando el vendedor fue eliminado
  active INTEGER NOT NULL DEFAULT 1,
  deleted INTEGER NOT NULL DEFAULT 0,
  owner_admin_id INTEGER,          -- admin que creó al vendedor (su dueño)
  owner_admin_name TEXT,           -- nombre del admin dueño (etiqueta visible)
  paid_cents INTEGER NOT NULL DEFAULT 0,  -- dinero que el vendedor ya entregó a su admin
  hidden INTEGER NOT NULL DEFAULT 0,      -- vendedor de INVITADOS: sus boletos no cuentan como venta
  commission_pct REAL,             -- su comisión propia (NULL = la general, 0 = ninguna)
  tutorial_seen INTEGER NOT NULL DEFAULT 0,  -- ya vio el tutorial de bienvenida
  es_lider INTEGER NOT NULL DEFAULT 0,       -- esta fila ES el colíder vendiendo en persona
  ultimo_ingreso TEXT,                       -- última vez que entró a la boletera
  created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS ticket_types (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  price_cents INTEGER NOT NULL,
  is_vip INTEGER NOT NULL DEFAULT 0,
  active INTEGER NOT NULL DEFAULT 1,
  needs_faculty INTEGER NOT NULL DEFAULT 1,  -- UADY la pide, VIP/Externo no
  flash_price_cents INTEGER        -- lo que cuesta cuando NO hay fase corriendo y se prende el flash
);
CREATE TABLE IF NOT EXISTS price_phases (
  -- fases de precio por tipo: al llegar la fecha de cada fase, el precio cambia solo
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  type_id INTEGER NOT NULL,
  name TEXT NOT NULL,              -- ej. Preventa, Fase 2, General
  price_cents INTEGER NOT NULL,
  starts_on TEXT NOT NULL,         -- fecha AAAA-MM-DD desde la que aplica
  group_pct INTEGER,               -- % de descuento de grupo de ESTA fase (NULL = usar el de Ajustes)
  es_flash INTEGER NOT NULL DEFAULT 0, -- venta flash: el boleto sale con el precio normal tachado
  flash_price_cents INTEGER        -- lo que cuesta ESTA fase cuando se prende el botón de flash
);
CREATE TABLE IF NOT EXISTS faculties (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,
  active INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS tickets (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  folio TEXT NOT NULL UNIQUE,
  qr_token TEXT NOT NULL UNIQUE,
  qr_payload TEXT,                 -- lo que va dentro del QR (token secreto)
  buyer_name TEXT NOT NULL,
  faculty_id INTEGER,
  faculty_name TEXT NOT NULL,      -- congelado al generar
  type_id INTEGER,
  type_name TEXT NOT NULL,         -- congelado al generar
  type_is_vip INTEGER NOT NULL DEFAULT 0,
  price_cents INTEGER NOT NULL,    -- congelado al generar (RF-40)
  seller_id INTEGER,
  seller_name TEXT NOT NULL,       -- congelado (se conserva si se elimina al vendedor)
  seller_code TEXT NOT NULL,
  status TEXT NOT NULL DEFAULT 'active',  -- active | used | void
  created_at TEXT NOT NULL,
  used_at TEXT,                    -- cuándo entró (primer escaneo en la puerta)
  voided_at TEXT,
  voided_by TEXT,
  void_reason TEXT,
  group_id INTEGER,             -- si el boleto es parte de un grupo (5 o 10), su id
  phase_name TEXT,               -- fase de precio vigente al generar (congelada), para el boleto
  group_size INTEGER,            -- 5 o 10 si es de grupo (congelado)
  es_representante INTEGER NOT NULL DEFAULT 0,  -- su boleto reclama la botella en barra
  es_cortesia INTEGER NOT NULL DEFAULT 0,       -- invitado especial: no dice precio, dice CORTESÍA
  normal_price_cents INTEGER,    -- precio individual antes del descuento de grupo (congelado, solo grupos)
  client_ref TEXT                -- id que manda la boletera para no duplicar si reintenta
);
CREATE TABLE IF NOT EXISTS groups (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  size INTEGER NOT NULL,             -- 5 o 10
  names TEXT NOT NULL,               -- JSON: lista de integrantes en orden
  representative TEXT,               -- solo en grupos de 10 (recibe la botella)
  seller_id INTEGER,
  seller_name TEXT NOT NULL,
  created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS sessions (
  token TEXT PRIMARY KEY,
  role TEXT NOT NULL,              -- seller | admin
  user_id INTEGER NOT NULL,
  created_at TEXT NOT NULL,
  expires_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS login_attempts (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  key TEXT NOT NULL,               -- ip o ip+usuario
  ts REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS audit_log (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  actor TEXT NOT NULL,
  action TEXT NOT NULL,
  detail TEXT NOT NULL,
  created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS seller_payments (
  -- cada entrega de dinero del vendedor a su admin. Se guarda el detalle completo
  -- para poder demostrarle después cómo fue pagando y cuánto se llevó de comisión.
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  seller_id INTEGER NOT NULL,
  seller_name TEXT NOT NULL,          -- congelado, se conserva si borran al vendedor
  amount_cents INTEGER NOT NULL,      -- abono: cuánto de su deuda se salda
  commission_cents INTEGER NOT NULL,  -- lo que el vendedor se queda de comisión
  cash_cents INTEGER NOT NULL,        -- efectivo que entregó = abono - comisión
  commission_pct REAL NOT NULL,       -- % con el que se calculó (congelado)
  note TEXT,
  created_by TEXT,                    -- admin que lo registró
  created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS expenses (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL,             -- ej. Local, Alcohol, DJ, Seguridad
  amount_cents INTEGER NOT NULL DEFAULT 0,
  account TEXT,                   -- quién lo paga (Russel, Osmar, Andry, o libre)
  status TEXT NOT NULL DEFAULT 'pendiente',   -- pendiente | pagado (se deriva de paid_cents)
  paid_cents INTEGER NOT NULL DEFAULT 0,      -- cuánto se ha abonado (los adelantos)
  created_by TEXT,
  created_at TEXT NOT NULL
);
"""

DEFAULT_SETTINGS = {
    "event_name": "HELLFIRE",
    "event_subtitle": "Noche de brujas",
    "event_date_text": "",
    "folio_prefix": "HF-",
    # Cierre de ventas: la noche de la fiesta se apaga la boletera para poder cortar
    # cuentas con los vendedores sabiendo que el número ya no se mueve. El ESCÁNER y
    # el panel de admin siguen funcionando: la puerta no puede depender de esto.
    "ventas_cerradas": "0",
    # Clave del escáner de la puerta. Vacía = el escáner solo funciona con sesión de
    # admin. El día del evento se genera una clave de 6 dígitos y se reparte al staff.
    "door_code": "",
    "folio_start": "1",              # número del primer folio (no revela lo vendido)
    "session_minutes": "480",
    "admin_session_minutes": "480",
    # La puerta aguanta toda la noche. Con las 8 h de los demás, quien abriera el
    # escáner a las 8 PM se quedaba fuera a las 4 AM, con gente formada enfrente.
    "scanner_session_minutes": "1080",   # 18 horas
    # Flyers por tipo de boleto: UADY, Externo, VIP, Grupo de 5 y Grupo de 10. Cada uno
    # con su imagen (base64 en la BD), posición y zoom. "gen" es el flyer "General" de
    # antes del cambio y sigue de respaldo (UADY/Externo caían ahí); las claves sin
    # sufijo son el flyer legado de una sola imagen, el último respaldo de todos.
    "flyer_file": "", "flyer_data": "", "flyer_mime": "",
    "flyer_focus": "0.5", "flyer_scale": "1",
    "flyer_data_vip": "", "flyer_mime_vip": "", "flyer_focus_vip": "", "flyer_scale_vip": "",
    "flyer_data_gen": "", "flyer_mime_gen": "", "flyer_focus_gen": "", "flyer_scale_gen": "",
    "flyer_data_uady": "", "flyer_mime_uady": "", "flyer_focus_uady": "", "flyer_scale_uady": "",
    "flyer_data_externo": "", "flyer_mime_externo": "", "flyer_focus_externo": "", "flyer_scale_externo": "",
    "flyer_data_grupo10": "", "flyer_mime_grupo10": "", "flyer_focus_grupo10": "", "flyer_scale_grupo10": "",
    # Ultra VIP: ya está a la venta como cualquier otro tipo
    "flyer_data_ultravip": "", "flyer_mime_ultravip": "", "flyer_focus_ultravip": "", "flyer_scale_ultravip": "",
    "flyer_data_grupo10vip": "", "flyer_mime_grupo10vip": "",
    "flyer_focus_grupo10vip": "", "flyer_scale_grupo10vip": "",
    "flyer_data_grupo10ultra": "", "flyer_mime_grupo10ultra": "",
    "flyer_focus_grupo10ultra": "", "flyer_scale_grupo10ultra": "",
    "flyer_data_cortesiaexterno": "", "flyer_mime_cortesiaexterno": "",
    "flyer_focus_cortesiaexterno": "", "flyer_scale_cortesiaexterno": "",
    "flyer_data_cortesiavip": "", "flyer_mime_cortesiavip": "",
    "flyer_focus_cortesiavip": "", "flyer_scale_cortesiavip": "",
    "flyer_data_cortesiaultra": "", "flyer_mime_cortesiaultra": "",
    "flyer_focus_cortesiaultra": "", "flyer_scale_cortesiaultra": "",
    # El interruptor de la venta flash. Con "1" hay flash AHORA, sin esperar fecha.
    "flash_manual": "0",
    "seller_commission_pct": "10",   # % de comisión del vendedor sobre lo que entrega
    "max_login_attempts": "8",
    "lockout_minutes": "10",
}

def setting(db, key):
    row = db.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else DEFAULT_SETTINGS.get(key, "")

def set_setting(db, key, value):
    db.execute("INSERT INTO settings(key,value) VALUES(?,?) "
               "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, str(value)))

FLYER_VARIANTS = ("uady", "externo", "vip", "grupo10", "ultravip",
                  "grupo10vip", "grupo10ultra",
                  "cortesiaexterno", "cortesiavip", "cortesiaultra")
FLYER_LABEL = {"uady": "UADY", "externo": "Externo", "vip": "VIP",
               "grupo10": "Grupo de 10", "ultravip": "Ultra VIP",
               "grupo10vip": "Grupo de 10 · VIP", "grupo10ultra": "Grupo de 10 · Ultra VIP",
               "cortesiaexterno": "Cortesía Externo", "cortesiavip": "Cortesía VIP",
               "cortesiaultra": "Cortesía Ultra VIP"}
# cadena de respaldo: si no han subido el flyer del tipo, usa el de un tipo
# relacionado antes de caer al flyer legado de una sola imagen
FLYER_FALLBACK = {"uady": "gen", "externo": "gen", "grupo10": "externo",
                   "ultravip": "vip",
                   # un grupo VIP sin flyer propio usa el de VIP, no el de grupo
                   # Externo: lo que el comprador reconoce primero es su categoría
                   "grupo10vip": "vip", "grupo10ultra": "ultravip",
                   # mientras no suban el suyo, la cortesía usa el flyer del tipo que
                   # le toca: se ve bien desde el primer invitado, sin configurar nada
                   "cortesiaexterno": "externo", "cortesiavip": "vip",
                   "cortesiaultra": "ultravip"}

def _flyer_chain(v):
    chain = [v]
    nxt = FLYER_FALLBACK.get(v)
    while nxt and nxt not in chain:
        chain.append(nxt)
        nxt = FLYER_FALLBACK.get(nxt)
    return chain

def flyer_info(db):
    """Configuración de los flyers (uno por tipo de boleto) para el frontend, con
    respaldo en cadena hasta el flyer legado de una sola imagen."""
    out = {}
    for v in FLYER_VARIANTS:
        data = None
        for vv in _flyer_chain(v):
            data = setting(db, f"flyer_data_{vv}")
            if data:
                break
        has = bool(data or setting(db, "flyer_data"))
        out[f"flyer_{v}"] = has
        out[f"flyer_focus_{v}"] = float(setting(db, f"flyer_focus_{v}")
                                        or setting(db, "flyer_focus") or 0.5)
        out[f"flyer_scale_{v}"] = float(setting(db, f"flyer_scale_{v}")
                                        or setting(db, "flyer_scale") or 1)
    return out

def flash_manual(db):
    """¿Está prendido el botón de venta flash? Manda sobre el calendario: con él
    encendido hay flash aunque ninguna fecha haya llegado; apagado, las fases flash
    del calendario siguen funcionando solas como respaldo."""
    return setting(db, "flash_manual") == "1"


def _nombre_base(nombre):
    """'Fase 2 Flash', 'FASE 2 flash', 'Fase 2' → 'fase 2'. Sirve para emparejar una
    venta flash del calendario con la fase normal a la que le hace el descuento."""
    n = (nombre or "").strip().lower()
    n = re.sub(r"[⚡]", " ", n)
    n = re.sub(r"\bflash\b", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def migrar_precios_flash(db):
    """Las ventas flash ya cargadas por fecha se convierten TAMBIÉN en el precio de
    flash de su fase ('Fase 2 Flash' $300 → Fase 2 vale $300 en flash), para que el
    botón sirva desde el primer día sin volver a teclear la tabla. No borra nada: las
    fases flash del calendario siguen ahí."""
    filas = db.execute("SELECT * FROM price_phases ORDER BY id").fetchall()
    porTipo = {}
    for f in filas:
        porTipo.setdefault(f["type_id"], []).append(f)
    tocados = 0
    for tid, fases in porTipo.items():
        flashes = [f for f in fases if f["es_flash"]]
        for f in flashes:
            base = _nombre_base(f["name"])
            for n in fases:
                if n["es_flash"] or n["flash_price_cents"] is not None:
                    continue
                if _nombre_base(n["name"]) == base and f["price_cents"] < n["price_cents"]:
                    db.execute("UPDATE price_phases SET flash_price_cents=? WHERE id=?",
                               (f["price_cents"], n["id"]))
                    tocados += 1
    if tocados:
        db.commit()
        print(f"[OnFire] Precios de venta flash copiados a su fase: {tocados}")


def effective_price(db, type_row):
    """Precio vigente de un tipo: la fase más reciente cuya fecha ya llegó; si no hay
    fase aplicable, el precio base. Devuelve (precio, nombre_fase, normal): normal
    solo viene cuando la fase vigente es una VENTA FLASH, y es el precio que regiría
    sin ella —el que el boleto saca tachado—. No se guarda a mano en ningún lado:
    se calcula ignorando las fases flash, así que nunca puede quedar desactualizado
    respecto a la fase real a la que se vuelve cuando el flash termina."""
    today = now_dt().strftime("%Y-%m-%d")
    ph = db.execute("""SELECT * FROM price_phases WHERE type_id=? AND starts_on<=?
                       ORDER BY starts_on DESC, id DESC LIMIT 1""",
                    (type_row["id"], today)).fetchone()
    if not ph:
        # Sin ninguna fase corriendo se vende al precio base. El botón TAMBIÉN tiene
        # que servir aquí: las fases pueden arrancar semanas después y mientras tanto
        # ya se está vendiendo. El precio de flash vive entonces en el tipo mismo.
        base = type_row["price_cents"]
        suyo = type_row["flash_price_cents"] if "flash_price_cents" in type_row.keys() else None
        if flash_manual(db) and suyo and suyo < base:
            return suyo, "Venta flash", base
        return base, None, None
    if not ph["es_flash"]:
        # EL INTERRUPTOR. Cada fase guarda su propio precio de flash, así que prender
        # el botón cobra el flash DE LA FASE QUE ESTÉ CORRIENDO: si se prende en
        # Fase 1 sale el de Fase 1, y si se vuelve a prender en Fase 4 sale el de
        # Fase 4, sin tocar nada. Apagar y volver a prender no cambia el precio.
        if flash_manual(db) and ph["flash_price_cents"] and ph["flash_price_cents"] < ph["price_cents"]:
            return ph["flash_price_cents"], ph["name"] + " Flash", ph["price_cents"]
        return ph["price_cents"], ph["name"], None
    # El tachado es el precio AL QUE SE VUELVE cuando el flash termine, o sea la
    # próxima fase normal. No el de la fase anterior: un flash suele ser la apertura
    # con descuento de su propia fase —"5 días antes de que suba a $425"—, y mirando
    # hacia atrás salía tachando un precio más barato que el que se está cobrando.
    sig = db.execute("""SELECT * FROM price_phases WHERE type_id=? AND starts_on>?
                        AND es_flash=0 ORDER BY starts_on ASC, id ASC LIMIT 1""",
                     (type_row["id"], today)).fetchone()
    if sig:
        normal = sig["price_cents"]
    else:
        # un flash sin fase que lo termine: se compara con lo último normal que hubo
        prev = db.execute("""SELECT * FROM price_phases WHERE type_id=? AND starts_on<=?
                             AND es_flash=0 ORDER BY starts_on DESC, id DESC LIMIT 1""",
                          (type_row["id"], today)).fetchone()
        normal = prev["price_cents"] if prev else type_row["price_cents"]
    if normal <= ph["price_cents"]:
        normal = None      # un "flash" más caro que lo normal no tacha nada
    return ph["price_cents"], ph["name"], normal

def avisos_flash(db, parsed, date, tipos):
    """Un flash que cuesta MÁS que la fase anterior no es una oferta: el que compró
    ayer pagó menos. El sistema no lo prohíbe —puede haber una razón— pero lo dice,
    porque en la pantalla los dos números se ven bien y el error solo aparece cuando
    un comprador reclama."""
    nombres = {t["id"]: t["name"] for t in tipos}
    avisos = []
    for tid, cents in parsed:
        previa = db.execute("""SELECT price_cents FROM price_phases WHERE type_id=?
                               AND starts_on<? AND es_flash=0
                               ORDER BY starts_on DESC, id DESC LIMIT 1""",
                            (tid, date)).fetchone()
        base = db.execute("SELECT price_cents FROM ticket_types WHERE id=?", (tid,)).fetchone()
        antes = previa["price_cents"] if previa else (base["price_cents"] if base else 0)
        if antes and cents > antes:
            avisos.append(f"{nombres.get(tid, tid)}: ${cents/100:,.0f} es MÁS caro que "
                          f"antes del flash (${antes/100:,.0f})")
        sig = db.execute("""SELECT price_cents FROM price_phases WHERE type_id=?
                            AND starts_on>? AND es_flash=0
                            ORDER BY starts_on ASC, id ASC LIMIT 1""",
                         (tid, date)).fetchone()
        if sig and cents >= sig["price_cents"]:
            avisos.append(f"{nombres.get(tid, tid)}: ${cents/100:,.0f} no baja de su fase "
                          f"(${sig['price_cents']/100:,.0f}), así que el boleto no llevará tachado")
    return avisos

def next_phase(db, type_row):
    """La próxima fase cuya fecha aún NO llegó (el siguiente cambio de precio).
    Devuelve {name, price_cents, starts_on} o None si no hay más fases futuras."""
    today = now_dt().strftime("%Y-%m-%d")
    ph = db.execute("""SELECT * FROM price_phases WHERE type_id=? AND starts_on>?
                       ORDER BY starts_on ASC, id ASC LIMIT 1""",
                    (type_row["id"], today)).fetchone()
    if ph:
        return {"name": ph["name"], "price_cents": ph["price_cents"],
                "starts_on": ph["starts_on"]}
    return None

def gen_seller_code(db):
    """Código de vendedor de 5 dígitos. Eran 4: con 50 vendedores, 1 de cada 200
    números al azar era un código válido y adivinar alguno tomaba ~3 horas desde un
    solo teléfono. Con 5 son 1 de cada 2,000 —más de un día por IP, y el candado de
    intentos avisa mucho antes—. Puro número a propósito: con letras se dicta mal y
    obliga a cambiar de teclado en el celular."""
    for _ in range(500):
        code = f"{secrets.randbelow(100000):05d}"
        taken = db.execute(
            "SELECT 1 FROM sellers WHERE code=? AND deleted=0", (code,)).fetchone()
        if not taken:
            return code
    raise RuntimeError("sin códigos disponibles")

def _schema_for_backend():
    s = SCHEMA
    if IS_PG:
        s = s.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
    return s

def borrar_todo(db, admin_principal, motivo):
    """Deja la base como recién estrenada. BORRA boletos, grupos, vendedores, gastos,
    fases, movimientos, sesiones y los demás admins (quien tenga sesión abierta
    pierde el acceso al instante). CONSERVA los flyers y ajustes (viven en settings),
    los precios, las facultades y el vendedor de invitados, que se recrea al arrancar
    desde GUEST_SELLER_CODE."""
    for table in ("tickets", "groups", "sellers", "price_phases",
                  "expenses", "audit_log", "login_attempts", "sessions"):
        db.execute(f"DELETE FROM {table}")
    db.execute("DELETE FROM admins WHERE username != ?", (admin_principal,))
    db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
               ("sistema", "inicializacion", motivo, now_iso()))
    db.commit()

def init_db():
    # espera a que la base esté disponible (Railway puede tardar unos segundos al arrancar)
    db = None
    for intento in range(10):
        try:
            db = db_connect()
            break
        except Exception as e:
            print(f"[OnFire] esperando la base de datos… ({e})")
            time.sleep(2)
    if db is None:
        raise RuntimeError("no se pudo conectar a la base de datos")

    db.executescript(_schema_for_backend())
    # migración suave: agregar columnas nuevas si la base viene de una versión anterior
    if IS_PG:
        for col in ("qr_payload TEXT", "used_at TEXT"):
            db.execute(f"ALTER TABLE tickets ADD COLUMN IF NOT EXISTS {col}")
        for col in ("owner_admin_id INTEGER", "owner_admin_name TEXT",
                    "paid_cents INTEGER NOT NULL DEFAULT 0",
                    "hidden INTEGER NOT NULL DEFAULT 0",
                    "commission_pct REAL",
                    "tutorial_seen INTEGER NOT NULL DEFAULT 0",
                    "es_lider INTEGER NOT NULL DEFAULT 0",
                    "ultimo_ingreso TEXT"):
            db.execute(f"ALTER TABLE sellers ADD COLUMN IF NOT EXISTS {col}")
        db.execute("ALTER TABLE admins ADD COLUMN IF NOT EXISTS "
                   "role TEXT NOT NULL DEFAULT 'admin'")
        db.execute("ALTER TABLE expenses ADD COLUMN IF NOT EXISTS "
                   "paid_cents INTEGER NOT NULL DEFAULT 0")
        db.execute("ALTER TABLE ticket_types ADD COLUMN IF NOT EXISTS "
                   "needs_faculty INTEGER NOT NULL DEFAULT 1")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS group_id INTEGER")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS phase_name TEXT")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS group_size INTEGER")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS "
                   "es_representante INTEGER NOT NULL DEFAULT 0")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS "
                   "es_cortesia INTEGER NOT NULL DEFAULT 0")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS normal_price_cents INTEGER")
        db.execute("ALTER TABLE price_phases ADD COLUMN IF NOT EXISTS group_pct INTEGER")
        db.execute("ALTER TABLE price_phases ADD COLUMN IF NOT EXISTS "
                   "es_flash INTEGER NOT NULL DEFAULT 0")
        db.execute("ALTER TABLE price_phases ADD COLUMN IF NOT EXISTS flash_price_cents INTEGER")
        db.execute("ALTER TABLE ticket_types ADD COLUMN IF NOT EXISTS flash_price_cents INTEGER")
        db.execute("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS client_ref TEXT")
    else:
        cols = [r["name"] for r in db.execute("PRAGMA table_info(tickets)").fetchall()]
        if "qr_payload" not in cols:
            db.execute("ALTER TABLE tickets ADD COLUMN qr_payload TEXT")
        if "used_at" not in cols:
            db.execute("ALTER TABLE tickets ADD COLUMN used_at TEXT")
        scols = [r["name"] for r in db.execute("PRAGMA table_info(sellers)").fetchall()]
        if "owner_admin_id" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN owner_admin_id INTEGER")
        if "owner_admin_name" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN owner_admin_name TEXT")
        if "paid_cents" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN paid_cents INTEGER NOT NULL DEFAULT 0")
        acols = [r["name"] for r in db.execute("PRAGMA table_info(admins)").fetchall()]
        if "role" not in acols:
            db.execute("ALTER TABLE admins ADD COLUMN role TEXT NOT NULL DEFAULT 'admin'")
        if "commission_pct" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN commission_pct REAL")
        if "tutorial_seen" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN tutorial_seen INTEGER NOT NULL DEFAULT 0")
        if "es_lider" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN es_lider INTEGER NOT NULL DEFAULT 0")
        if "ultimo_ingreso" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN ultimo_ingreso TEXT")
        gcols = [r["name"] for r in db.execute("PRAGMA table_info(expenses)").fetchall()]
        if "paid_cents" not in gcols:
            db.execute("ALTER TABLE expenses ADD COLUMN paid_cents INTEGER NOT NULL DEFAULT 0")
        if "hidden" not in scols:
            db.execute("ALTER TABLE sellers ADD COLUMN hidden INTEGER NOT NULL DEFAULT 0")
        ttcols = [r["name"] for r in db.execute("PRAGMA table_info(ticket_types)").fetchall()]
        if "needs_faculty" not in ttcols:
            db.execute("ALTER TABLE ticket_types ADD COLUMN needs_faculty INTEGER NOT NULL DEFAULT 1")
        tkcols = [r["name"] for r in db.execute("PRAGMA table_info(tickets)").fetchall()]
        if "group_id" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN group_id INTEGER")
        if "phase_name" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN phase_name TEXT")
        if "group_size" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN group_size INTEGER")
        if "normal_price_cents" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN normal_price_cents INTEGER")
        if "client_ref" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN client_ref TEXT")
        if "es_representante" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN es_representante INTEGER NOT NULL DEFAULT 0")
        if "es_cortesia" not in tkcols:
            db.execute("ALTER TABLE tickets ADD COLUMN es_cortesia INTEGER NOT NULL DEFAULT 0")
        pcols = [r["name"] for r in db.execute("PRAGMA table_info(price_phases)").fetchall()]
        if "group_pct" not in pcols:
            db.execute("ALTER TABLE price_phases ADD COLUMN group_pct INTEGER")
        if "es_flash" not in pcols:
            db.execute("ALTER TABLE price_phases ADD COLUMN es_flash INTEGER NOT NULL DEFAULT 0")
        if "flash_price_cents" not in pcols:
            db.execute("ALTER TABLE price_phases ADD COLUMN flash_price_cents INTEGER")
        tcols = [r["name"] for r in db.execute("PRAGMA table_info(ticket_types)").fetchall()]
        if "flash_price_cents" not in tcols:
            db.execute("ALTER TABLE ticket_types ADD COLUMN flash_price_cents INTEGER")
    db.commit()
    for k, v in DEFAULT_SETTINGS.items():
        db.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO NOTHING", (k, v))
    db.commit()

    migrar_precios_flash(db)

    # los invitados de antes de la marca se reconocen por su serie de folio (INV-)
    db.execute("UPDATE tickets SET es_cortesia=1 WHERE es_cortesia=0 AND folio LIKE ?",
               ("INV-%",))
    db.commit()

    # corrección de una sola vez: ningún pago puede exceder lo vendido (datos viejos)
    if setting(db, "paid_capped_v1") != "1":
        sub = ("SELECT COALESCE(SUM(CASE WHEN status!='void' THEN price_cents ELSE 0 END),0) "
               "FROM tickets WHERE seller_id = sellers.id")
        db.execute(f"UPDATE sellers SET paid_cents = ({sub}) WHERE paid_cents > ({sub})")
        set_setting(db, "paid_capped_v1", "1")
        db.commit()

    # limpieza de una sola vez: el admin ya no genera boletos; se borran los que creó
    if setting(db, "drop_admin_tickets_v1") != "1":
        db.execute("DELETE FROM tickets WHERE seller_id IS NULL AND seller_name LIKE ?",
                   ("Admin: %",))
        set_setting(db, "drop_admin_tickets_v1", "1")
        db.commit()

    # Admin inicial: si defines ADMIN_USER + ADMIN_PASSWORD (en Railway → Variables),
    # el admin arranca con TUS credenciales. Si no, usa el admin por defecto solo en local.
    env_user = (os.environ.get("ADMIN_USER") or "").strip()
    env_pass = os.environ.get("ADMIN_PASSWORD") or ""
    use_env = bool(env_user and env_pass)
    init_user = env_user if use_env else "admin"
    init_pass = env_pass if use_env else "onfire2026"

    n_admins = db.execute("SELECT COUNT(*) AS c FROM admins").fetchone()["c"]
    if n_admins == 0:
        # primera vez: crear admin inicial + catálogo + 4 vendedores
        db.execute("INSERT INTO admins(username, pass_hash, created_at) VALUES(?,?,?)",
                   (init_user, hash_password(init_pass), now_iso()))
        # 3 precios (UADY, Externo, VIP) en 2 tipos: UADY/Externo son "General" (no VIP);
        # UADY pide facultad, Externo y VIP no. Arrancan en $0: el admin los define
        # en Catálogos antes de iniciar (no se puede vender con precio en 0).
        for name, price, vip, needs in [("UADY", 0, 0, 1), ("Externo", 0, 0, 0),
                                        ("VIP", 0, 1, 0)]:
            db.execute("INSERT INTO ticket_types(name, price_cents, is_vip, needs_faculty) "
                       "VALUES(?,?,?,?)", (name, price, vip, needs))
        for f in ["Ingeniería", "Medicina", "Derecho", "Economía", "Arquitectura"]:
            db.execute("INSERT INTO faculties(name) VALUES(?)", (f,))
        # RF-25: el sistema inicia con 4 códigos activos, uno por vendedor
        codes = []
        for i in range(1, 5):
            code = None
            while code is None:
                c = f"{secrets.randbelow(100000):05d}"
                if not db.execute("SELECT 1 FROM sellers WHERE code=?", (c,)).fetchone():
                    code = c
            codes.append(code)
            db.execute("INSERT INTO sellers(name, code, created_at) VALUES(?,?,?)",
                       (f"Vendedor {i}", code, now_iso()))
        db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
                   ("sistema", "inicializacion",
                    f"Sistema inicializado. Admin inicial: {init_user}", now_iso()))
        db.commit()
        if not use_env:   # solo guardamos la contraseña en archivo cuando es la de por defecto (local)
            try:
                with open(os.path.join(DATA, "CREDENCIALES_INICIALES.txt"), "w") as f:
                    f.write("OnFire — credenciales iniciales\n================================\n\n")
                    f.write(f"Administrador:  usuario: {init_user}   contraseña: {init_pass}\n\n")
                    for i, c in enumerate(codes, 1):
                        f.write(f"Vendedor {i}: código {c}\n")
            except OSError:
                pass
        print(f"[OnFire] Base creada. Admin: '{init_user}'"
              + ("" if use_env else f"/{init_pass}")
              + f" · Códigos vendedor: {', '.join(codes)}")
    elif use_env and not db.execute("SELECT 1 FROM admins WHERE username=?", (env_user,)).fetchone():
        # ya había datos, pero definiste un admin por variables que aún no existía → crearlo
        db.execute("INSERT INTO admins(username, pass_hash, created_at) VALUES(?,?,?)",
                   (env_user, hash_password(env_pass), now_iso()))
        db.commit()
        print(f"[OnFire] Admin '{env_user}' creado desde ADMIN_USER/ADMIN_PASSWORD.")

    # Limpieza TOTAL de una sola vez (evento nuevo): borra todo lo generado en las
    # pruebas — boletos, vendedores, fases, movimientos, sesiones y los demás admins.
    # Solo queda el admin inicial (ADMIN_USER/ADMIN_PASSWORD) y los catálogos
    # (tipos de boleto y facultades). Corre una vez y se marca con la bandera.
    if setting(db, "event_reset_v1") != "1":
        for table in ("tickets", "sellers", "price_phases", "audit_log",
                      "login_attempts", "sessions"):
            db.execute(f"DELETE FROM {table}")
        db.execute("DELETE FROM admins WHERE username != ?", (init_user,))
        db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
                   ("sistema", "inicializacion",
                    f"Evento reiniciado: datos de prueba eliminados. Admin inicial: {init_user}",
                    now_iso()))
        set_setting(db, "event_reset_v1", "1")
        db.commit()
        print(f"[OnFire] Limpieza total: solo queda el admin '{init_user}'.")

    # Reestructura de una sola vez: 2 tipos "General" (UADY/Externo) + VIP, 3 precios.
    # UADY pide facultad; Externo y VIP no. 'Externo' deja de ser facultad.
    if setting(db, "types_uady_externo_v1") != "1":
        db.execute("UPDATE ticket_types SET name='UADY', needs_faculty=1 WHERE name='General'")
        db.execute("UPDATE ticket_types SET needs_faculty=0 WHERE is_vip=1")  # VIP sin facultad
        if not db.execute("SELECT 1 FROM ticket_types WHERE name='Externo'").fetchone():
            # precio inicial del Externo = el de UADY (ajústalo en Catálogos)
            base = db.execute("SELECT price_cents FROM ticket_types WHERE name='UADY'").fetchone()
            db.execute("INSERT INTO ticket_types(name, price_cents, is_vip, needs_faculty) "
                       "VALUES(?,?,?,?)", ("Externo", (base["price_cents"] if base else 25000), 0, 0))
        else:
            db.execute("UPDATE ticket_types SET is_vip=0, needs_faculty=0 WHERE name='Externo'")
        db.execute("DELETE FROM faculties WHERE name='Externo'")
        set_setting(db, "types_uady_externo_v1", "1")
        db.commit()
        print("[OnFire] Tipos reestructurados: UADY / Externo / VIP.")

    # RESET para lanzamiento: borra TODO lo generado (boletos, vendedores, fases,
    # movimientos, sesiones y demás admins) y deja los precios en $0 para que el
    # admin los defina antes de iniciar. Solo queda el admin inicial y los catálogos.
    if setting(db, "event_reset_v2") != "1":
        for table in ("tickets", "sellers", "price_phases", "audit_log",
                      "login_attempts", "sessions"):
            db.execute(f"DELETE FROM {table}")
        db.execute("DELETE FROM admins WHERE username != ?", (init_user,))
        db.execute("UPDATE ticket_types SET price_cents=0")   # precios en 0: editar para iniciar
        db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
                   ("sistema", "inicializacion",
                    f"Sistema listo para lanzar: datos borrados, precios en 0. Admin: {init_user}",
                    now_iso()))
        set_setting(db, "event_reset_v2", "1")
        db.commit()
        print(f"[OnFire] RESET de lanzamiento: solo admin '{init_user}', precios en 0.")

    def limpiar_todo(motivo):
        borrar_todo(db, init_user, motivo)

    # RESET v3 — el borrón de una sola vez que se pidió antes del lanzamiento.
    if setting(db, "event_reset_v3") != "1":
        limpiar_todo("Sistema reiniciado para el lanzamiento: boletos, vendedores, "
                     f"gastos y movimientos borrados. Sesiones cerradas. Admin: {init_user}")
        set_setting(db, "event_reset_v3", "1")
        db.commit()
        print(f"[OnFire] RESET v3: base limpia, unico admin '{init_user}'.")

    # La facultad es SOLO de los boletos UADY: ni Externo, ni VIP, ni Ultra VIP
    # pertenecen a una. Los tipos creados antes nacieron con la casilla marcada
    # (venía activa por defecto), así que la pedían y la imprimían en el boleto.
    #
    # El primer intento solo miraba los marcados como VIP ★, y un "Ultra VIP" creado
    # sin esa marca se quedaba pidiendo facultad igual. Aquí se corrige por lo que
    # de verdad importa: el único tipo con facultad es UADY. Corre una vez; si algún
    # día se quiere otra cosa, se marca desde Editar y esto ya no vuelve a tocarlo.
    # No es una migración de una sola vez: se revisa en CADA arranque. La versión
    # anterior se marcaba como hecha en el primer arranque —cuando todavía no existía
    # el tipo problemático— y después ya nunca corregía nada. Un tipo creado más tarde
    # con la casilla mal se quedaba pidiendo facultad para siempre.
    #
    # Lo único que respeta es la decisión explícita de un admin: si alguien marca a
    # mano "Pedir facultad" desde Editar, ese tipo queda en la lista de excepciones
    # (facultad_manual) y esto no lo vuelve a tocar.
    manual = set(json.loads(setting(db, "facultad_manual") or "[]"))
    arreglados = [r for r in db.execute(
        "SELECT id, name FROM ticket_types WHERE needs_faculty=1 "
        "AND LOWER(TRIM(name)) != 'uady'").fetchall() if r["id"] not in manual]
    if arreglados:
        for r in arreglados:
            db.execute("UPDATE ticket_types SET needs_faculty=0 WHERE id=?", (r["id"],))
        nombres = ", ".join(r["name"] for r in arreglados)
        db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
                   ("sistema", "catalogo",
                    f"Ya no piden facultad (solo UADY la lleva): {nombres}", now_iso()))
        db.commit()
        print(f"[OnFire] Sin facultad: {nombres}")

    # Códigos de vendedor a 5 dígitos. Se revisa en CADA arranque (no es bandera de
    # una sola vez): cualquier vendedor visible con código corto recibe uno nuevo.
    # El de invitados NO se toca —viene de la variable de entorno y es privado del
    # organizador—. Las sesiones abiertas siguen vivas: al vendedor no se le corta
    # la venta, solo necesita el código nuevo para su próximo login.
    cortos = db.execute(
        "SELECT id, name FROM sellers WHERE deleted=0 AND hidden=0 "
        "AND code IS NOT NULL AND LENGTH(code) < 5").fetchall()
    if cortos:
        for r in cortos:
            db.execute("UPDATE sellers SET code=? WHERE id=?", (gen_seller_code(db), r["id"]))
        db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
                   ("sistema", "usuarios",
                    f"Códigos de vendedor renovados a 5 dígitos ({len(cortos)}): "
                    "compárteles el nuevo desde Vendedores", now_iso()))
        db.commit()
        print(f"[OnFire] {len(cortos)} códigos de vendedor renovados a 5 dígitos.")

    # RESET A PETICIÓN — para volver a dejar el sistema en cero cuantas veces haga
    # falta (por ejemplo después de enseñárselo al equipo), sin tocar código:
    # en Railway → Variables, pon RESET_KEY con cualquier valor. Al arrancar borra
    # todo y guarda ese valor; mientras NO lo cambies no vuelve a borrar, así que un
    # reinicio normal del servicio jamás se lleva los datos por accidente. Para
    # reiniciar otra vez, cambia el valor (ej. "demo1" -> "demo2").
    reset_key = (os.environ.get("RESET_KEY") or "").strip()
    if reset_key and setting(db, "reset_key_aplicada") != reset_key:
        limpiar_todo(f"Sistema reiniciado a peticion (RESET_KEY). Admin: {init_user}")
        set_setting(db, "reset_key_aplicada", reset_key)
        db.commit()
        print(f"[OnFire] RESET a peticion ({reset_key}): base limpia.")

    # VENDEDOR DE INVITADOS (oculto). Va AL FINAL, después de todos los resets, para
    # que ninguna limpieza lo borre. Su código vive en la variable de entorno
    # GUEST_SELLER_CODE, no en la base: si algún día se borra o se recrea la base, al
    # arrancar se vuelve a crear el MISMO código y los boletos de invitados siguen
    # escaneando. Sus boletos NO cuentan como venta en ninguna pantalla del panel.
    guest_code = (os.environ.get("GUEST_SELLER_CODE") or "").strip()
    if guest_code:
        if not re.fullmatch(r"\d{4,6}", guest_code):
            print(f"[OnFire] AVISO: GUEST_SELLER_CODE='{guest_code}' no sirve. "
                  "Debe ser de 4 a 6 dígitos (ej. 482113). No se creó el vendedor de invitados.")
        else:
            row = db.execute("SELECT * FROM sellers WHERE code=?", (guest_code,)).fetchone()
            if row and not row["hidden"]:
                # ese código ya es de un vendedor REAL: no lo tocamos. Convertirlo en
                # oculto borraría sus ventas del resumen sin que nadie se diera cuenta.
                print(f"[OnFire] AVISO: el código {guest_code} ya es del vendedor "
                      f"'{row['name']}'. Elige otro GUEST_SELLER_CODE; no se creó "
                      "el vendedor de invitados.")
            elif row:   # ya es el de invitados: asegurar que siga utilizable
                if not row["active"] or row["deleted"]:
                    db.execute("UPDATE sellers SET active=1, deleted=0 WHERE id=?", (row["id"],))
                    db.commit()
                print("[OnFire] Vendedor de invitados (oculto) listo.")
            else:
                # Cambiar GUEST_SELLER_CODE tiene que APAGAR el código anterior. Antes se
                # creaba un segundo vendedor de invitados y el viejo seguía funcionando:
                # el código se cambia justo cuando se filtró, así que dejarlo vivo anula
                # el motivo del cambio. Se renombra el que ya existe —no se borra, sus
                # boletos siguen colgando de él— y se cierran sus sesiones abiertas.
                viejos = db.execute(
                    "SELECT * FROM sellers WHERE hidden=1 AND deleted=0 ORDER BY id").fetchall()
                if viejos:
                    principal = viejos[0]
                    db.execute("UPDATE sellers SET code=?, active=1 WHERE id=?",
                               (guest_code, principal["id"]))
                    db.execute("DELETE FROM sessions WHERE role='seller' AND user_id=?",
                               (principal["id"],))
                    for extra in viejos[1:]:
                        # duplicados de versiones anteriores: sin código no se puede
                        # entrar con ellos, pero sus boletos se conservan
                        db.execute("UPDATE sellers SET code=NULL, active=0 WHERE id=?",
                                   (extra["id"],))
                        db.execute("DELETE FROM sessions WHERE role='seller' AND user_id=?",
                                   (extra["id"],))
                    db.commit()
                    print(f"[OnFire] Código de invitados actualizado; el anterior quedó "
                          f"apagado ({len(viejos) - 1} duplicado(s) desactivado(s)).")
                else:
                    db.execute("INSERT INTO sellers(name, code, hidden, created_at) VALUES(?,?,1,?)",
                               (os.environ.get("GUEST_SELLER_NAME") or "Invitados",
                                guest_code, now_iso()))
                    db.commit()
                    print("[OnFire] Vendedor de invitados (oculto) creado.")

    db.commit()
    db.close()

# ---------------------------------------------------------------- Excel (sincronización automática)

HEADERS = ["Folio", "Comprador", "Facultad", "Tipo de boleto", "Precio",
           "Vendedor", "Código vendedor", "Fecha de venta", "Estado",
           "Ingresó", "Hora de ingreso", "Anulado por", "Motivo anulación"]

STATUS_ES = {"active": "ACTIVO", "used": "INGRESÓ", "void": "ANULADO"}

def _ticket_row(t):
    return [
        t["folio"], t["buyer_name"], t["faculty_name"], t["type_name"],
        money(t["price_cents"]), t["seller_name"], t["seller_code"],
        t["created_at"], STATUS_ES.get(t["status"], t["status"]),
        "Sí" if t["used_at"] else "No", t["used_at"] or "",
        t["voided_by"] or "", t["void_reason"] or "",
    ]

def build_workbook(rows, summary=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Boletos"
    header_fill = PatternFill("solid", fgColor="1F1005")
    header_font = Font(bold=True, color="FF8A3D")
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(horizontal="center")
    for t in rows:
        ws.append(_ticket_row(t))
    for i, w in enumerate([10, 28, 16, 14, 10, 22, 14, 19, 10, 9, 19, 16, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in row:
            c.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"
    if summary is not None:
        ws2 = wb.create_sheet("Resumen por vendedor")
        ws2.append(["Vendedor", "Admin", "Código", "Boletos válidos", "Boletos anulados",
                    "Monto total", "Pagado", "Estatus"])
        for c in ws2[1]:
            c.fill, c.font = header_fill, header_font
        for s in summary:
            paid = s.get("paid_cents") or 0
            settled = s["total_cents"] > 0 and paid >= s["total_cents"]
            ws2.append([s["name"], s.get("owner_admin_name") or "—", s["code"] or "—",
                        s["count_valid"], s["count_void"], money(s["total_cents"]),
                        money(paid), "COMPLETADO" if settled else "Pendiente"])
        for i, w in enumerate([24, 16, 10, 15, 16, 14, 12, 13], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        for row in ws2.iter_rows(min_row=2, min_col=6, max_col=7):
            for c in row:
                c.number_format = '"$"#,##0.00'
    return wb

def seller_summary(db):
    return [dict(r) for r in db.execute("""
        SELECT s.name, s.code, s.owner_admin_name, s.paid_cents,
          COALESCE(SUM(CASE WHEN t.status!='void' THEN 1 ELSE 0 END),0) AS count_valid,
          COALESCE(SUM(CASE WHEN t.status='void' THEN 1 ELSE 0 END),0)  AS count_void,
          COALESCE(SUM(CASE WHEN t.status!='void' THEN t.price_cents ELSE 0 END),0) AS total_cents
        FROM sellers s LEFT JOIN tickets t ON t.seller_id=s.id
        WHERE s.hidden=0
        GROUP BY s.id ORDER BY total_cents DESC""").fetchall()]

def sync_excel():
    """Regenera boletos.xlsx con todas las ventas.
    (Es un archivo derivado; la fuente de verdad es la base de datos.)"""
    try:
        db = db_connect()
        rows = db.execute(
            f"SELECT * FROM tickets WHERE {NOT_GUEST} ORDER BY id").fetchall()
        summary = seller_summary(db)
        db.close()
        wb = build_workbook(rows, summary)
        # el temporal lleva el id del hilo: si dos sincronizaciones se cruzan, cada
        # una escribe el suyo y ninguna se queda sin archivo al renombrar
        tmp = f"{XLSX}.{threading.get_ident()}.tmp"
        wb.save(tmp)
        os.replace(tmp, XLSX)
    except Exception as e:
        print(f"[OnFire] error al sincronizar Excel: {e}")

# El Excel se regenera ENTERO cada vez, así que hacerlo en cada venta es trabajo al
# cuadrado: con 1500 boletos serían 1500 reconstrucciones de un archivo que no para
# de crecer, y en plena noche de venta eso ahoga al servidor. En vez de eso se marca
# "hay cambios" y UN SOLO hilo lo regenera cada pocos segundos. Al ser uno solo,
# tampoco pueden chocar dos escrituras sobre el mismo archivo.
_excel_pendiente = threading.Event()
SYNC_CADA = 8   # segundos entre regeneraciones, como mucho

def _excel_worker():
    while True:
        _excel_pendiente.wait()      # dormido hasta que haya algo que guardar
        _excel_pendiente.clear()
        sync_excel()
        time.sleep(SYNC_CADA)        # agrupa las ventas de estos segundos en una sola

def sync_excel_async():
    _excel_pendiente.set()

# ---------------------------------------------------------------- respaldos (RG-04)

def backup_loop():
    while True:
        try:
            stamp = now_dt().strftime("%Y%m%d_%H%M")
            if not IS_PG and os.path.exists(DB_PATH):   # con Postgres el respaldo lo gestiona la plataforma
                shutil.copy2(DB_PATH, os.path.join(BACKUPS, f"onfire_{stamp}.db"))
            if os.path.exists(XLSX):
                shutil.copy2(XLSX, os.path.join(BACKUPS, f"boletos_{stamp}.xlsx"))
            keep = sorted(os.listdir(BACKUPS))
            for old in keep[:-40]:
                os.remove(os.path.join(BACKUPS, old))
        except Exception as e:
            print(f"[OnFire] error de respaldo: {e}")
        time.sleep(600)  # cada 10 minutos

# ---------------------------------------------------------------- auth / sesiones

def audit(db, actor, action, detail):
    db.execute("INSERT INTO audit_log(actor, action, detail, created_at) VALUES(?,?,?,?)",
               (actor, action, detail, now_iso()))

# Los boletos de INVITADOS (vendedor oculto) no son una venta: se excluyen del
# resumen, la cobranza, el listado de boletos y los movimientos. Siguen
# existiendo como fila normal en tickets, que es lo que el escáner necesita para
# validarlos en la puerta.
NOT_GUEST = "seller_id NOT IN (SELECT id FROM sellers WHERE hidden=1)"

def is_guest_seller(seller_row):
    try:
        return bool(seller_row["hidden"])
    except (KeyError, IndexError, TypeError):
        return False

def ticket_is_guest(db, t):
    """True si el boleto lo generó el vendedor de invitados. Se usa para que los
    endpoints que buscan un boleto POR ID (que es consecutivo y por tanto se puede
    ir adivinando) respondan 'no existe' a los admins."""
    if t is None or t["seller_id"] is None:
        return False
    row = db.execute("SELECT hidden FROM sellers WHERE id=?", (t["seller_id"],)).fetchone()
    return bool(row and row["hidden"])

# DOBLE TOQUE del código de invitados: el primer intento se rechaza con EXACTAMENTE
# el mismo mensaje que un código inválido, y el segundo (dentro de la ventana) sí
# entra. Así, quien ande probando claves al azar cree que falló y sigue de largo.
# Vive en memoria: gunicorn corre con --workers 1, así que todos los hilos la comparten.
# Si el servicio reinicia se olvida, y lo único que pasa es tocar dos veces otra vez.
GUEST_KNOCK_SECONDS = 180
_guest_knock = {}
_knock_lock = threading.Lock()

def guest_knock_ok(ip):
    """False en el primer intento (hay que repetir el código); True en el segundo."""
    now = time.time()
    with _knock_lock:
        for k, t in list(_guest_knock.items()):   # limpia toques viejos
            if now - t > GUEST_KNOCK_SECONDS:
                _guest_knock.pop(k, None)
        if _guest_knock.get(ip) is None:
            _guest_knock[ip] = now
            return False
        _guest_knock.pop(ip, None)
        return True

def create_session(db, role, user_id):
    minutes = int(setting(db, {"admin": "admin_session_minutes",
                              "scanner": "scanner_session_minutes"}.get(role, "session_minutes")))
    token = secrets.token_urlsafe(24)
    exp = (now_dt() + timedelta(minutes=minutes)).strftime("%Y-%m-%d %H:%M:%S")
    db.execute("INSERT INTO sessions(token, role, user_id, created_at, expires_at) VALUES(?,?,?,?,?)",
               (token, role, user_id, now_iso(), exp))
    return token

def client_ip():
    """La IP real del cliente. Detrás del proxy de Railway, remote_addr es la IP del
    PROXY —la misma para todo el mundo—, así que usarla para el candado de intentos
    convertía 8 códigos mal tecleados por cualquiera en un bloqueo para todos."""
    return (request.headers.get("X-Forwarded-For", request.remote_addr or "?")
            .split(",")[0].strip())

def current_session():
    token = (request.headers.get("Authorization") or "").replace("Bearer ", "").strip()
    if not token:
        return None
    db = get_db()
    s = db.execute("SELECT * FROM sessions WHERE token=?", (token,)).fetchone()
    if not s or s["expires_at"] < now_iso():
        return None
    if s["role"] == "seller":
        seller = db.execute("SELECT * FROM sellers WHERE id=?", (s["user_id"],)).fetchone()
        # RF-32 / RF-86: código desactivado o vendedor eliminado → sesión inválida
        if not seller or not seller["active"] or seller["deleted"]:
            db.execute("DELETE FROM sessions WHERE token=?", (token,))
            db.commit()
            return None
        return {"role": "seller", "seller": seller, "token": token}
    if s["role"] == "scanner":
        # sesión de puerta: no apunta a ningún usuario, solo autoriza escanear.
        # Si el admin apagó o rotó la clave, la sesión ya se borró y no llega aquí.
        return {"role": "scanner", "token": token}
    admin = db.execute("SELECT * FROM admins WHERE id=?", (s["user_id"],)).fetchone()
    if not admin:
        return None
    return {"role": "admin", "admin": admin, "token": token}

def require_seller():
    s = current_session()
    if not s or s["role"] != "seller":
        return None
    return s

def es_colider(s):
    return bool(s) and s["role"] == "admin" and (s["admin"]["role"] or "admin") == "colider"

def require_admin():
    """SOLO administradores de verdad. El colíder entra por require_panel().

    Deliberadamente NO se le abre aquí: así, cualquier ruta que exista hoy o que se
    agregue mañana queda cerrada para el colíder mientras nadie decida lo contrario.
    Una lista de permisos se olvida; una de prohibiciones se olvida y se filtra."""
    s = current_session()
    if not s or s["role"] != "admin":
        return None
    if es_colider(s):
        return None
    return s

def require_panel():
    """Admin o colíder. Solo para lo que el colíder SÍ puede: ver su grupo, dar de
    alta a sus vendedores y cobrarles. Quien la use tiene que filtrar por dueño."""
    s = current_session()
    if not s or s["role"] != "admin":
        return None
    return s

def mi_ambito(s):
    """El id del dueño cuyos datos puede ver quien está en sesión. None = todo el
    sistema (admin). Un número = solo lo de ese colíder."""
    return s["admin"]["id"] if es_colider(s) else None

def rate_limited(db, key):
    max_tries = int(setting(db, "max_login_attempts"))
    window = int(setting(db, "lockout_minutes")) * 60
    cutoff = time.time() - window
    db.execute("DELETE FROM login_attempts WHERE ts < ?", (cutoff,))
    n = db.execute("SELECT COUNT(*) c FROM login_attempts WHERE key=? AND ts>=?",
                   (key, cutoff)).fetchone()["c"]
    return n >= max_tries

def record_attempt(db, key):
    db.execute("INSERT INTO login_attempts(key, ts) VALUES(?,?)", (key, time.time()))

def clear_attempts(db, key):
    db.execute("DELETE FROM login_attempts WHERE key=?", (key,))

# ---------------------------------------------------------------- API: acceso

@app.get("/api/event")
def public_event():
    """Solo nombre/subtítulo del evento para las pantallas de acceso (sin datos personales)."""
    db = get_db()
    return jsonify(event_name=setting(db, "event_name"),
                   event_subtitle=setting(db, "event_subtitle"))

@app.post("/api/login-code")
def login_code():
    db = get_db()
    ip = client_ip()
    key = f"code:{ip}"
    if rate_limited(db, key):
        return jsonify(error="Demasiados intentos. Espera unos minutos."), 429
    # RF-28: mensaje genérico. El código de invitados usa ESTE MISMO texto en su
    # primer intento; si fuera distinto, delataría que ese código sí existe.
    BAD = "Código incorrecto. Intenta de nuevo."
    code = str((request.json or {}).get("code", "")).strip()
    # 4 a 6 dígitos: los códigos nuevos son de 5; el de invitados (variable de
    # entorno) puede ser de 4 a 6 y no hay por qué delatarlo rechazándolo
    if not re.fullmatch(r"\d{4,6}", code):
        record_attempt(db, key); db.commit()
        return jsonify(error=BAD), 401
    seller = db.execute(
        "SELECT * FROM sellers WHERE code=? AND active=1 AND deleted=0", (code,)).fetchone()
    if not seller:
        record_attempt(db, key); db.commit()
        return jsonify(error=BAD), 401
    # El doble toque del código de invitados se quitó. Existía cuando sus boletos eran
    # invisibles: si alguien adivinaba el código, los boletos gratis no aparecían en
    # ningún lado. Ahora salen en Boletos, en el filtro Cortesía y en su apartado, y se
    # pueden anular uno por uno — el problema se ve y se corrige, así que la molestia de
    # teclearlo dos veces cada vez ya no compra nada.
    clear_attempts(db, key)
    token = create_session(db, "seller", seller["id"])
    # Se marca cada entrada. Sirve para lo que no se puede saber de otro modo: quién
    # de verdad está trabajando y quién nada más quería el boleto. Sin vender no se
    # distingue "no ha podido" de "ni abrió la app"; con esto sí.
    db.execute("UPDATE sellers SET ultimo_ingreso=? WHERE id=?", (now_iso(), seller["id"]))
    db.commit()
    # La primera vez que entra se le enseña el tutorial. Se guarda en el SERVIDOR y no
    # en el teléfono: si cambia de celular o borra los datos del navegador ya lo vio,
    # y volvérselo a poner sería tratarlo como si no supiera vender.
    primera = not seller["tutorial_seen"] and not seller["hidden"]
    return jsonify(token=token, name=seller["name"], first_time=primera)

@app.post("/api/admin/login")
def admin_login():
    db = get_db()
    ip = client_ip()
    body = request.json or {}
    username = str(body.get("username", "")).strip()
    key = f"admin:{ip}"
    if rate_limited(db, key):
        return jsonify(error="Demasiados intentos. Espera unos minutos."), 429
    admin = db.execute("SELECT * FROM admins WHERE username=?", (username,)).fetchone()
    if not admin or not check_password(str(body.get("password", "")), admin["pass_hash"]):
        record_attempt(db, key); db.commit()
        return jsonify(error="Usuario o contraseña incorrectos"), 401
    clear_attempts(db, key)
    token = create_session(db, "admin", admin["id"])
    db.commit()
    return jsonify(token=token, username=admin["username"])

@app.post("/api/logout")
def logout():
    s = current_session()
    if s:
        db = get_db()
        db.execute("DELETE FROM sessions WHERE token=?", (s["token"],))
        db.commit()
    return jsonify(ok=True)

@app.get("/api/me")
def me():
    s = current_session()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    info = {"event_name": setting(db, "event_name"),
            "event_subtitle": setting(db, "event_subtitle"),
            "event_date_text": setting(db, "event_date_text"),
            "flyer": bool(setting(db, "flyer_file"))}
    if s["role"] == "seller":
        return jsonify(role="seller", name=s["seller"]["name"], **info)
    return jsonify(role="admin", name=s["admin"]["username"],
                   admin_id=s["admin"]["id"], **info)

def puede_gestionar(db, admin, sel):
    """¿Este admin puede editar / desactivar / eliminar a este vendedor?

    Igual que con anular y con cobrar: "cada admin con lo suyo" vale entre pares,
    no hacia abajo. Sin esta excepción, el equipo de un colíder —y el colíder
    mismo— no los podría tocar NADIE: el colíder porque no llega a estas rutas,
    y el admin porque el dueño figuraba como otro."""
    return owns_seller(admin, sel) or duenio_es_colider(db, sel)

def puede_cobrar(db, sesion, sel):
    """¿Quién le registra pagos a este vendedor?

    Dos agujeros que solo se ven poniendo dinero de por medio:

    - La ficha personal del colíder es suya, así que owns_seller() le decía que sí:
      podía marcarse a sí mismo como pagado y dejar su deuda en cero sin entregar un
      peso. Su propia cuenta se la cobra un admin, nunca él.
    - Y al revés: como esa ficha es suya, al admin le decía que no. Nadie podía
      cobrarle al colíder. La regla de "cada admin con lo suyo" es entre pares; hacia
      abajo el admin cobra.
    """
    admin = sesion["admin"]
    if es_colider(sesion):
        if sel["owner_admin_id"] != admin["id"]:
            return False, "no es de tu grupo"
        if sel["es_lider"]:
            return False, "Tu propia cuenta te la cobra un administrador"
        return True, None
    if owns_seller(admin, sel) or duenio_es_colider(db, sel):
        return True, None
    return False, f"Solo {sel['owner_admin_name']} (su admin) puede registrar pagos de este vendedor"

def es_admin_principal(admin):
    """El dueño del evento: el usuario de la variable de entorno."""
    return admin["username"] == (os.environ.get("ADMIN_USER") or "admin").strip()

def duenio_es_colider(db, seller_row):
    """¿Este vendedor pertenece al equipo de un colíder? Entonces sus datos son
    visibles hacia arriba: el colíder rinde cuentas al admin."""
    if not seller_row["owner_admin_id"]:
        return False
    a = db.execute("SELECT role FROM admins WHERE id=?",
                   (seller_row["owner_admin_id"],)).fetchone()
    return bool(a) and (a["role"] or "admin") == "colider"

def owns_seller(admin, seller_row):
    """Un admin es dueño del vendedor si lo creó. Vendedores antiguos sin dueño
    (owner NULL, de versiones previas) pueden gestionarse por cualquier admin."""
    return seller_row["owner_admin_id"] is None or seller_row["owner_admin_id"] == admin["id"]

# ---------------------------------------------------------------- API: vendedor

@app.get("/api/catalog")
def catalog():
    s = current_session()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    types = []
    for r in db.execute("SELECT * FROM ticket_types WHERE active=1 ORDER BY price_cents").fetchall():
        price, phase, normal = effective_price(db, r)
        types.append({"id": r["id"], "name": r["name"], "is_vip": r["is_vip"],
                      "needs_faculty": r["needs_faculty"],
                      "price_cents": price, "phase": phase,
                      "normal_cents": normal,
                      "next_phase": next_phase(db, r)})
    facs = [dict(r) for r in db.execute(
        "SELECT id, name FROM faculties WHERE active=1 ORDER BY name").fetchall()]
    # Plan de grupo (solo de 10, solo boletos Externo). YA NO LLEVA DESCUENTO: el
    # beneficio del grupo es la botella del representante, no el precio. Lo que antes
    # era el descuento ahora se le paga al vendedor como comisión.
    # Los tipos que pueden ir en grupo: todos los que no piden facultad. Se manda la
    # lista completa para que la boletera ofrezca Externo, VIP y Ultra VIP.
    opciones = [t for t in types if not t["needs_faculty"] and t["price_cents"] > 0]
    externo = next((t for t in opciones if t["name"] == "Externo"), None)
    group_info = None
    if opciones:
        base = externo or opciones[0]
        group_info = {"type_id": base["id"], "pct": 0,
                      "normal_price_cents": base["price_cents"],
                      "group_price_cents": base["price_cents"],
                      "savings_cents": 0,
                      "tipos": [{"id": t["id"], "name": t["name"], "is_vip": t["is_vip"],
                                 "price_cents": t["price_cents"],
                                 "normal_cents": t.get("normal_cents")} for t in opciones]}
    # ¿le falta el tutorial? Va aquí y no solo en la respuesta del login: si el
    # vendedor recarga la página a media guía, con la sesión ya guardada no vuelve a
    # pasar por el login y se quedaría sin verla nunca.
    pendiente = False
    if s.get("seller"):
        r = db.execute("SELECT tutorial_seen, hidden FROM sellers WHERE id=?",
                       (s["seller"]["id"],)).fetchone()
        pendiente = bool(r and not r["tutorial_seen"] and not r["hidden"])
    return jsonify(types=types, faculties=facs, group=group_info,
                   ventas_cerradas=ventas_cerradas(db),
                   # una flash prendida a mano no tiene hora de fin: el panel del
                   # vendedor no puede prometer un cronómetro que no existe
                   flash_manual=flash_manual(db),
                   tutorial_pendiente=pendiente,
                   event_name=setting(db, "event_name"),
                   event_subtitle=setting(db, "event_subtitle"),
                   event_date_text=setting(db, "event_date_text"),
                   **flyer_info(db))

def ventas_cerradas(db):
    return setting(db, "ventas_cerradas") == "1"

def ticket_public(t):
    return {"id": t["id"], "folio": t["folio"], "qr_token": t["qr_token"],
            "qr_payload": t["qr_payload"] or t["qr_token"],   # lo que va dentro del QR
            "buyer_name": t["buyer_name"], "faculty_name": t["faculty_name"],
            "type_name": t["type_name"], "type_is_vip": t["type_is_vip"],
            "price": money(t["price_cents"]), "status": t["status"],
            "created_at": t["created_at"], "used_at": t["used_at"],
            "seller_name": t["seller_name"], "seller_code": t["seller_code"],
            "phase_name": t["phase_name"], "group_size": t["group_size"],
            "es_representante": bool(t["es_representante"]),
            "es_cortesia": bool(t["es_cortesia"]),
            "normal_price": money(t["normal_price_cents"]) if t["normal_price_cents"] else None}

def _insert_ticket_row(db, buyer, fac_id, fac_name, tt, price_cents,
                        seller_id, seller_name, seller_code, group_id=None,
                        phase_name=None, group_size=None, normal_price_cents=None,
                        guest=False, client_ref=None, representante=False):
    # guest ya venía marcando la serie de folios (INV-); ahora también congela la
    # cortesía en el boleto, para que el flyer y el precio no dependan de ir a
    # buscar de qué vendedor salió
    """Inserta un boleto con folio único (reintenta si choca) y devuelve la fila.
    Compartido por la generación individual y la generación de grupos.

    Los boletos de invitado llevan su PROPIA serie de folios (INV-). Si tomaran
    números de la serie de venta, en el Excel quedarían huecos (…0001, 0003…) y
    eso delataría que hubo boletos que nadie puede ver. El folio no se imprime en
    el boleto ni lo muestra el escáner, así que el prefijo distinto no se nota."""
    prefix = "INV-" if guest else setting(db, "folio_prefix")
    with _write_lock:
        # el máximo se busca SOLO dentro de la misma serie, para que las dos
        # numeraciones (venta e invitados) avancen sin estorbarse
        base = db.execute(
            f"SELECT COALESCE(MAX(CAST(SUBSTR(folio, ?) AS INTEGER)),0) AS n "
            f"FROM tickets WHERE folio {LIKE} ?",
            (len(prefix) + 1, prefix + "%")).fetchone()["n"]
        if not guest:   # el folio inicial configurable solo aplica a la venta
            try:
                base = max(base, int(setting(db, "folio_start") or 1) - 1)
            except ValueError:
                pass
        for attempt in range(20):
            n = base + 1 + attempt
            folio = f"{prefix}{n:04d}"
            token = secrets.token_urlsafe(12)   # RF-46: no adivinable ni secuencial
            qr_payload = token
            try:
                cur = db.execute("""INSERT INTO tickets
                    (folio, qr_token, qr_payload, buyer_name, faculty_id, faculty_name,
                     type_id, type_name, type_is_vip, price_cents,
                     seller_id, seller_name, seller_code, status, created_at, group_id,
                     phase_name, group_size, normal_price_cents, client_ref,
                     es_representante, es_cortesia)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?, 'active', ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (folio, token, qr_payload, buyer, fac_id, fac_name, tt["id"], tt["name"],
                     tt["is_vip"], price_cents, seller_id, seller_name,
                     seller_code, now_iso(), group_id, phase_name, group_size,
                     normal_price_cents, client_ref,
                     1 if representante else 0, 1 if guest else 0))
                db.commit()
                break
            except IntegrityError:
                db.rollback()   # Postgres: liberar la transacción abortada antes de reintentar
                continue
        else:
            return None
    return db.execute("SELECT * FROM tickets WHERE id=?", (cur.lastrowid,)).fetchone()

@app.post("/api/tutorial-visto")
def tutorial_visto():
    """El vendedor terminó el tutorial de bienvenida. No se vuelve a mostrar."""
    s = require_seller()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    db.execute("UPDATE sellers SET tutorial_seen=1 WHERE id=?", (s["seller"]["id"],))
    db.commit()
    return jsonify(ok=True)

@app.post("/api/tickets")
def create_ticket():
    s = require_seller()   # solo los vendedores generan boletos (el admin ya no)
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # Ventas cerradas: la noche de la fiesta se apaga para cortar cuentas sabiendo que
    # el número ya no se mueve. El código de invitados es la excepción a propósito:
    # es privado del organizador y sirve de salida para una cortesía de último momento.
    if ventas_cerradas(db) and not is_guest_seller(s["seller"]):
        return jsonify(error="Las ventas ya cerraron. Habla con tu administrador."), 403
    body = request.json or {}
    # Si al vendedor se le cayó el internet DESPUÉS de que el boleto se creó, no
    # recibió respuesta y va a volver a darle a Generar. Sin esto quedaban dos
    # boletos de la misma venta y su cuenta salía cobrando de más. La boletera manda
    # la misma referencia en el reintento, así que aquí devolvemos el que ya existe.
    ref = str(body.get("client_ref", "")).strip()[:64]
    if ref:
        ya = db.execute("SELECT * FROM tickets WHERE client_ref=? AND seller_id=?",
                        (ref, s["seller"]["id"])).fetchone()
        if ya:
            return jsonify(ticket=ticket_public(ya), repetido=True)
    buyer = str(body.get("buyer_name", "")).strip()
    if len(buyer) < 3:
        return jsonify(error="Escribe el nombre completo del comprador"), 400
    tt = db.execute("SELECT * FROM ticket_types WHERE id=? AND active=1",
                    (body.get("type_id"),)).fetchone()
    if not tt:
        return jsonify(error="Elige un tipo de boleto válido"), 400
    # la facultad solo se pide para tipos que la requieren (UADY); Externo y VIP no
    if tt["needs_faculty"]:
        fac = db.execute("SELECT * FROM faculties WHERE id=? AND active=1",
                         (body.get("faculty_id"),)).fetchone()
        if not fac:
            return jsonify(error="Elige una facultad válida"), 400
        fac_id, fac_name = fac["id"], fac["name"]
    else:
        fac_id, fac_name = None, ""
    price_now, phase_name, normal_now = effective_price(db, tt)   # congelado en el boleto
    if price_now <= 0:   # el sistema no vende hasta que el admin defina el precio
        return jsonify(error="El precio de este boleto aún no está configurado. "
                             "Pídele al administrador que lo defina en Catálogos."), 400
    # RF-43: el boleto queda ligado al vendedor que lo genera
    seller_id, seller_name, seller_code = s["seller"]["id"], s["seller"]["name"], s["seller"]["code"]
    # en venta flash el boleto congela también el precio normal, para salir tachado
    t = _insert_ticket_row(db, buyer, fac_id, fac_name, tt, price_now,
                           seller_id, seller_name, seller_code, phase_name=phase_name,
                           normal_price_cents=normal_now,
                           guest=is_guest_seller(s["seller"]), client_ref=ref or None)
    if not t:
        return jsonify(error="No se pudo generar el folio, intenta de nuevo"), 500
    if not is_guest_seller(s["seller"]):   # los invitados no dejan rastro en Movimientos
        audit(db, seller_name, "generacion",
              f"Generó el boleto {t['folio']} para {buyer} ({tt['name']})")
    db.commit()
    sync_excel_async()
    return jsonify(ticket=ticket_public(t))

# ---- grupos de 5 y 10 (plan aparte, solo boletos Externo con descuento) --------

@app.post("/api/groups")
def create_group():
    """Genera un grupo de 10: cada integrante recibe su propio boleto del MISMO tipo,
    y uno de ellos —el representante— se lleva la botella.

    El tipo lo elige el vendedor (Externo, VIP o Ultra VIP): los grupos se piden en
    las tres categorías y antes salía siempre Externo, así que un grupo VIP se
    vendía como VIP y se generaba como general. No se pueden mezclar tipos dentro
    de un grupo: eso volvería ambiguo qué botella le toca a quién."""
    s = require_seller()
    if not s:
        return jsonify(error="sin sesión"), 401
    # los grupos salen en su propia pestaña del panel: el vendedor de invitados no
    # los usa (sus boletos se generan uno por uno y no deben aparecer en ningún lado)
    if is_guest_seller(s["seller"]):
        return jsonify(error="Esta cuenta genera boletos de invitado uno por uno"), 403
    db = get_db()
    if ventas_cerradas(db):
        return jsonify(error="Las ventas ya cerraron. Habla con tu administrador."), 403
    b = request.json or {}
    size = b.get("size")
    if size != 10:
        return jsonify(error="El grupo debe ser de exactamente 10 integrantes"), 400
    names = b.get("names") or []
    if not isinstance(names, list) or len(names) != size:
        return jsonify(error=f"Escribe los {size} nombres del grupo"), 400
    names = [str(n).strip() for n in names]
    for n in names:
        if len(n) < 3:
            return jsonify(error="Cada integrante necesita su nombre completo"), 400
    representative = None
    if size == 10:
        idx = b.get("representative_index")
        if not isinstance(idx, int) or idx < 0 or idx >= 10:
            return jsonify(error="Marca quién es el representante del grupo (recibe la botella)"), 400
        representative = names[idx]
    # Sin type_id se asume Externo: así los grupos que ya existían siguen igual.
    tid = b.get("type_id")
    if tid:
        tt = db.execute("SELECT * FROM ticket_types WHERE id=? AND active=1", (tid,)).fetchone()
    else:
        tt = db.execute("SELECT * FROM ticket_types WHERE name='Externo' AND active=1").fetchone()
    if not tt:
        return jsonify(error="Ese tipo de boleto no está disponible para armar grupos"), 400
    # La facultad es por persona y aquí solo se piden los diez nombres: un grupo UADY
    # saldría con la facultad vacía en los diez boletos.
    if tt["needs_faculty"]:
        return jsonify(error=f"Los grupos no se pueden armar con {tt['name']}, "
                             f"porque cada boleto necesita su facultad."), 400
    price_now, phase_name, normal_now = effective_price(db, tt)
    if price_now <= 0:
        return jsonify(error=f"El precio de {tt['name']} aún no está configurado"), 400
    pct = 0   # el grupo ya no tiene descuento: su beneficio es la botella
    group_price = round(price_now * (100 - pct) / 100)
    seller_id, seller_name, seller_code = s["seller"]["id"], s["seller"]["name"], s["seller"]["code"]
    gcur = db.execute("INSERT INTO groups(size, names, representative, seller_id, seller_name, created_at) "
                      "VALUES(?,?,?,?,?,?)",
                      (size, json.dumps(names, ensure_ascii=False), representative,
                       seller_id, seller_name, now_iso()))
    db.commit()
    gid = gcur.lastrowid
    tickets_out = []
    for name in names:
        # La marca va en SU boleto, no solo en el registro del grupo: el de la barra
        # no tiene el panel abierto, tiene un boleto enfrente. Si los diez se ven
        # iguales, cualquiera puede decir que él es el representante.
        t = _insert_ticket_row(db, name, None, "", tt, group_price,
                               seller_id, seller_name, seller_code, group_id=gid,
                               phase_name=phase_name, group_size=size,
                               # en venta flash el grupo también saca su tachado
                               normal_price_cents=normal_now,
                               representante=(name == representative))
        if not t:
            return jsonify(error="No se pudo generar uno de los folios, intenta de nuevo"), 500
        tickets_out.append(ticket_public(t))
    audit(db, seller_name, "generacion",
          f"Generó un grupo de {size} ({', '.join(names)}) a ${group_price/100:,.2f} c/u"
          + (f" · representante: {representative}" if representative else ""))
    db.commit()
    sync_excel_async()
    return jsonify(group_id=gid, size=size, representative=representative,
                   price=money(group_price), normal_price=money(price_now),
                   savings=money(price_now - group_price), tickets=tickets_out)

@app.get("/api/admin/groups")
def list_groups():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    rows = db.execute("""
        SELECT g.*, s.owner_admin_name FROM groups g
        LEFT JOIN sellers s ON s.id = g.seller_id
        ORDER BY g.id DESC LIMIT 300""").fetchall()
    out = []
    for r in rows:
        folios = [dict(t) for t in db.execute(
            "SELECT folio, status FROM tickets WHERE group_id=? ORDER BY id", (r["id"],)).fetchall()]
        out.append({"id": r["id"], "size": r["size"], "names": json.loads(r["names"]),
                    "representative": r["representative"], "seller_name": r["seller_name"],
                    "owner_admin_name": r["owner_admin_name"],
                    "created_at": r["created_at"], "tickets": folios})
    return jsonify(groups=out)

@app.get("/api/my-tickets")
def my_tickets():
    s = require_seller()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    q = (request.args.get("q") or "").strip()
    sql = "SELECT * FROM tickets WHERE seller_id=?"
    params = [s["seller"]["id"]]
    if q:
        sql += f" AND (buyer_name {LIKE} ? OR folio {LIKE} ?)"
        params += [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY id DESC"   # RF-72
    rows = db.execute(sql, params).fetchall()
    count = db.execute(
        "SELECT COUNT(*) c FROM tickets WHERE seller_id=? AND status!='void'",
        (s["seller"]["id"],)).fetchone()["c"]   # RF-55/68: anulados no cuentan
    return jsonify(count=count, tickets=[ticket_public(t) for t in rows])

@app.get("/api/tickets/<int:tid>")
def get_ticket(tid):
    s = current_session()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    t = db.execute("SELECT * FROM tickets WHERE id=?", (tid,)).fetchone()
    if not t:
        return jsonify(error="no existe"), 404
    if s["role"] == "seller" and t["seller_id"] != s["seller"]["id"]:
        return jsonify(error="no existe"), 404   # RF-74: nunca boletos de otro
    # para un admin, los boletos de invitado no existen (los ids son consecutivos,
    # así que sin esto bastaría con irlos probando uno por uno para verlos)
    if s["role"] == "admin" and ticket_is_guest(db, t):
        return jsonify(error="no existe"), 404
    return jsonify(ticket=ticket_public(t))

# ---------------------------------------------------------------- API: escaneo en la puerta

@app.get("/api/scan/recent")
def scan_recent():
    """Los últimos que entraron. Lo pide el propio escáner para que el staff pueda
    contestar "¿ya pasó fulano?" sin llamar al organizador, y para saber cuántos van
    adentro. Solo nombre, tipo y hora: nada de precios ni de quién lo vendió."""
    if not require_scanner():
        return jsonify(error="clave requerida"), 401
    db = get_db()
    filas = db.execute(
        f"""SELECT buyer_name, type_name, type_is_vip, used_at FROM tickets
            WHERE status='used' AND {NOT_GUEST}
            ORDER BY used_at DESC LIMIT 60""").fetchall()
    total = db.execute(
        f"SELECT COUNT(*) c FROM tickets WHERE status='used' AND {NOT_GUEST}").fetchone()["c"]
    return jsonify(total=total, entradas=[dict(r) for r in filas])

@app.post("/api/scan-login")
def scan_login():
    """Entrada del staff de la puerta: la clave de 6 dígitos que el organizador
    genera el día del evento. Con el mismo candado de intentos que los demás logins."""
    db = get_db()
    ip = client_ip()
    if rate_limited(db, "door:" + ip):
        return jsonify(error="Demasiados intentos. Espera unos minutos."), 429
    codigo = str((request.json or {}).get("code", "")).strip()
    puerta = setting(db, "door_code")
    if not puerta or codigo != puerta:
        db.execute("INSERT INTO login_attempts(key, ts) VALUES(?,?)", ("door:" + ip, time.time()))
        db.commit()
        return jsonify(error="Clave incorrecta"), 401
    token = create_session(db, "scanner", 0)
    db.commit()
    return jsonify(token=token)

def require_scanner():
    """Puede escanear: un admin (siempre) o una sesión de puerta (con la clave).
    El colíder NO por ser colíder: escanear quema el boleto, y eso es irreversible.
    Si el día del evento se para en la puerta, se le pasa la clave del staff como a
    cualquier otro."""
    s = current_session()
    if not s or s["role"] not in ("admin", "scanner"):
        return None
    if es_colider(s):
        return None
    return s

@app.post("/api/scan")
def scan():
    """Valida un boleto en tiempo real y lo marca como INGRESÓ en el primer escaneo.
    Cierra el boleto: cualquier copia o falso sale en rojo.

    YA NO es público. Lo era —"sin un boleto real no se puede hacer nada"— pero eso
    ignoraba dos cosas: acepta FOLIOS, que son consecutivos y adivinables, y hasta con
    puros tokens cualquiera con la URL podía quemar el boleto de otra persona con una
    foto. Ahora escanea el admin siempre, y el staff con la clave del día del evento."""
    if not require_scanner():
        return jsonify(error="clave requerida"), 401
    db = get_db()
    ident = folio_from_scan((request.json or {}).get("code", ""))
    if not ident:
        return jsonify(result="no_existe")
    t = db.execute("SELECT * FROM tickets WHERE qr_token=? OR folio=?",
                   (ident, ident.upper())).fetchone()
    if not t:
        return jsonify(result="no_existe")   # QR falso / folio inexistente
    if t["status"] == "void":
        return jsonify(result="anulado", ticket=ticket_public(t))
    if t["status"] == "used":
        return jsonify(result="usado", used_at=t["used_at"], ticket=ticket_public(t))
    # Primer escaneo → marcar ingreso. El WHERE status='active' hace que solo UNA de
    # dos peticiones simultáneas cambie la fila; el desempate es cuántas filas tocó
    # ESTA petición. Antes se comparaba used_at, pero esa marca tiene precisión de
    # SEGUNDOS: dos copias escaneadas en el mismo segundo daban la misma hora, las
    # dos creían haber ganado y ENTRABAN LAS DOS.
    when = now_iso()
    cur = db.execute("UPDATE tickets SET status='used', used_at=? WHERE id=? AND status='active'",
                     (when, t["id"]))
    gane = cur.rowcount == 1
    db.commit()
    t2 = db.execute("SELECT * FROM tickets WHERE id=?", (t["id"],)).fetchone()
    if not gane:   # otro escáner llegó primero por milésimas
        return jsonify(result="usado", used_at=t2["used_at"], ticket=ticket_public(t2))
    sync_excel_async()
    return jsonify(result="valido", ticket=ticket_public(t2))

# ---------------------------------------------------------------- API: administrador

@app.get("/api/admin/summary")
def admin_summary():
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # El colíder ve SOLO lo de su grupo. No es que se le oculte el total: es que para
    # él ese total no existe, porque nunca recibe una fila que no sea suya.
    duenio = mi_ambito(s)
    filtro = (" AND seller_id IN (SELECT id FROM sellers WHERE owner_admin_id=?)"
              if duenio else "")
    par = (duenio,) if duenio else ()
    tot = db.execute(f"""SELECT
        SUM(CASE WHEN status!='void' THEN 1 ELSE 0 END) AS n,
        SUM(CASE WHEN status!='void' THEN price_cents ELSE 0 END) AS cents,
        SUM(CASE WHEN status='used' THEN 1 ELSE 0 END) AS entered
        FROM tickets WHERE {NOT_GUEST}{filtro}""", par).fetchone()
    paid = db.execute(
        "SELECT COALESCE(SUM(paid_cents),0) AS c FROM sellers WHERE hidden=0" +
        (" AND owner_admin_id=?" if duenio else ""), par).fetchone()["c"]
    # desglose por admin: cuánto han vendido sus vendedores y cuánto ya cobró (todos lo ven)
    by_admin = db.execute(f"""
        SELECT COALESCE(s.owner_admin_name, 'Sin asignar') AS admin_name,
               COALESCE(SUM(s.paid_cents),0) AS paid_cents,
               COALESCE(SUM(tk.sold),0) AS sold_cents
        FROM sellers s
        LEFT JOIN (SELECT seller_id, SUM(CASE WHEN status!='void' THEN price_cents ELSE 0 END) AS sold
                   FROM tickets GROUP BY seller_id) tk ON tk.seller_id = s.id
        WHERE s.deleted=0 AND s.hidden=0{" AND s.owner_admin_id=?" if duenio else ""}
        GROUP BY COALESCE(s.owner_admin_name, 'Sin asignar')
        ORDER BY sold_cents DESC""", par).fetchall()
    admins = [{"admin": r["admin_name"], "sold": money(r["sold_cents"]),
               "collected": money(r["paid_cents"]),
               "settled": r["sold_cents"] > 0 and r["paid_cents"] >= r["sold_cents"]}
              for r in by_admin]
    return jsonify(total_tickets=tot["n"] or 0, total=money(tot["cents"] or 0),
                   entered=tot["entered"] or 0, collected=money(paid), by_admin=admins,
                   soy_colider=bool(duenio), yo=s["admin"]["username"])

def ticket_filters(prefix="", con_cortesias=False):
    """WHERE dinámico compartido por la tabla admin y la exportación (RF-93).
    prefix: alias de la tabla tickets cuando la consulta usa JOIN (ej. "t.").

    con_cortesias: los boletos de invitado siguen fuera de todo por defecto —esa
    regla es la que los hace invisibles—, pero la pestaña Boletos del dueño ahora
    los muestra a propósito, con "Cortesía" donde va el precio. La exportación NO
    los lleva: ese archivo es el registro del dinero, y un invitado no pagó."""
    a = request.args
    p = prefix
    where, params = [f"({p}{NOT_GUEST}" + (f" OR {p}es_cortesia=1)" if con_cortesias else ")")], []
    if a.get("admin"):   # boletos de un admin: los de SUS vendedores + los que él generó
        if a["admin"] == "__none__":
            where.append(f"{p}seller_id IN (SELECT id FROM sellers WHERE owner_admin_name IS NULL)")
        else:
            where.append(
                f"({p}seller_id IN (SELECT id FROM sellers WHERE owner_admin_name=?) "
                f"OR {p}seller_name=?)")
            params.append(a["admin"])
            params.append("Admin: " + a["admin"])
    if a.get("seller_id"):
        where.append(f"{p}seller_id=?"); params.append(a["seller_id"])
    if a.get("faculty"):
        where.append(f"{p}faculty_name=?"); params.append(a["faculty"])
    if a.get("type"):
        # "Cortesía" no es un tipo de boleto, es una forma de entrar: el invitado
        # puede ser VIP o Ultra VIP. Pero en el filtro se busca como si lo fuera,
        # porque es como se piensa al buscarlo: "enséñame las cortesías".
        if a["type"] == "__cortesia__":
            where.append(f"{p}es_cortesia=1")
        else:
            where.append(f"{p}type_name=? AND {p}es_cortesia=0"); params.append(a["type"])
    if a.get("q"):
        where.append(f"({p}buyer_name {LIKE} ? OR {p}folio {LIKE} ?)")
        params += [f"%{a['q']}%", f"%{a['q']}%"]
    return (" WHERE " + " AND ".join(where) if where else ""), params

@app.get("/api/admin/tickets")
def admin_tickets():
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    duenio = mi_ambito(s)
    where, params = ticket_filters("t.", con_cortesias=not duenio)
    if duenio:
        # el colíder ve boletos, pero solo los de su grupo — y las cortesías son del
        # dueño del evento, así que ni de reojo (con_cortesias ya lo dejó fuera)
        where += (" AND " if where else " WHERE ") + "s.owner_admin_id=?"
        params = list(params) + [duenio]
    rows = db.execute(
        "SELECT t.*, s.owner_admin_id AS owner_admin_id, s.owner_admin_name AS owner_admin_name "
        "FROM tickets t LEFT JOIN sellers s ON s.id = t.seller_id"
        + where + " ORDER BY t.id DESC", params).fetchall()
    me = s["admin"]
    colideres = {r["id"] for r in db.execute(
        "SELECT id FROM admins WHERE role='colider'").fetchall()}
    out = []
    for t in rows:
        tp = ticket_public(t)
        tp["owner_admin_id"] = t["owner_admin_id"]
        tp["owner_admin_name"] = t["owner_admin_name"]
        # el SERVIDOR decide si este admin puede anular este boleto (única verdad)
        if t["seller_id"] is not None:
            cv = (t["owner_admin_id"] is None or t["owner_admin_id"] == me["id"]
                  or (t["owner_admin_id"] in colideres))
        else:   # boleto generado por un admin → solo ese mismo admin
            creator = (t["seller_name"] or "").replace("Admin: ", "", 1)
            cv = creator == me["username"]
        # anular es de admins. Al colíder ni se le pinta el botón.
        tp["can_void"] = cv and not es_colider(s)
        out.append(tp)
    return jsonify(tickets=out)

def can_void(admin, db, t):
    """Solo el admin dueño del vendedor puede anular sus boletos. Boletos generados
    directamente por un admin: solo ese mismo admin. Vendedores sin dueño (legado):
    cualquier admin."""
    if t["seller_id"] is not None:
        sel = db.execute("SELECT * FROM sellers WHERE id=?", (t["seller_id"],)).fetchone()
        if sel and sel["owner_admin_id"] is not None and sel["owner_admin_id"] != admin["id"]:
            # Los del equipo de un colíder son la excepción: él tiene prohibido anular,
            # así que si el admin tampoco pudiera, esos boletos no los podría cancelar
            # NADIE. La regla de "cada admin con lo suyo" es entre pares, no hacia abajo.
            if not duenio_es_colider(db, sel):
                return False, sel["owner_admin_name"]
        return True, None
    # boleto generado por un admin (seller_name = "Admin: usuario")
    creator = (t["seller_name"] or "").removeprefix("Admin: ")
    if creator and creator != admin["username"]:
        return False, creator
    return True, None

@app.post("/api/admin/tickets/<int:tid>/void")
def void_ticket(tid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    reason = str((request.json or {}).get("reason", "")).strip()
    if not reason:
        return jsonify(error="Escribe el motivo de la anulación"), 400
    t = db.execute("SELECT * FROM tickets WHERE id=?", (tid,)).fetchone()
    # Las cortesías SÍ se anulan: con 100 repartidas, que una se filtre o que alguien
    # falle es cuestión de tiempo, y el boleto tiene que poder cancelarse. Los demás
    # boletos de invitado (los que no son cortesía) siguen sin existir para el admin.
    if not t or (ticket_is_guest(db, t) and not t["es_cortesia"]):
        return jsonify(error="no existe"), 404
    if t["status"] == "void":
        return jsonify(error="Ya estaba anulado"), 400
    ok, owner = can_void(s["admin"], db, t)
    if not ok:
        return jsonify(error=f"Solo {owner} (admin del vendedor) puede anular este boleto"), 403
    db.execute("UPDATE tickets SET status='void', voided_at=?, voided_by=?, void_reason=? WHERE id=?",
               (now_iso(), s["admin"]["username"], reason, tid))
    audit(db, s["admin"]["username"], "anulacion",
          f"Anuló el boleto {t['folio']} de {t['buyer_name']} ({t['type_name']}, "
          f"vendió {t['seller_name']}). Motivo: {reason}")
    db.commit()
    sync_excel_async()
    return jsonify(ok=True)

# ---- catálogos: tipos de boleto y facultades (RF-80/81)

@app.get("/api/admin/ticket-types")
def list_types():
    s = require_panel()   # LEER el catálogo también el colíder: sin esto,
    # la pestaña Boletos truena al cargar sus filtros y el 401 lo saca
    # del panel con un "tu sesión terminó" que no explica nada.
    # Cambiarlo (PUT/POST/DELETE) sigue siendo solo de admins.
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    out = []
    for r in db.execute("SELECT * FROM ticket_types ORDER BY id").fetchall():
        price, phase, _n = effective_price(db, r)
        phases = [dict(p) for p in db.execute(
            "SELECT * FROM price_phases WHERE type_id=? ORDER BY starts_on, id",
            (r["id"],)).fetchall()]
        vendidos = db.execute(
            "SELECT COUNT(*) c FROM tickets WHERE type_id=? AND status!='void'",
            (r["id"],)).fetchone()["c"]
        out.append({**dict(r), "current_price_cents": price,
                    "current_phase": phase, "phases": phases,
                    "sold": vendidos})
    return jsonify(types=out)

@app.post("/api/admin/ticket-types/<int:tid>/phases")
def create_phase(tid):
    """Nueva fase: nombre, precio y fecha. Al llegar la fecha, el precio cambia solo."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    t = db.execute("SELECT * FROM ticket_types WHERE id=?", (tid,)).fetchone()
    if not t:
        return jsonify(error="no existe"), 404
    b = request.json or {}
    name = str(b.get("name", "")).strip()
    date = str(b.get("starts_on", "")).strip()
    try:
        price = int(round(float(b.get("price", 0)) * 100))
    except (TypeError, ValueError):
        price = 0
    if not name or price <= 0 or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
        return jsonify(error="Fase incompleta: nombre, precio y fecha (AAAA-MM-DD)"), 400
    # es_flash se ignoraba aquí: una fase creada por esta vía cobraba el precio de
    # oferta pero NO tachaba nada, así que el descuento no se veía por ningún lado.
    es_flash = 1 if b.get("es_flash") else 0
    db.execute("INSERT INTO price_phases(type_id, name, price_cents, starts_on, es_flash) "
               "VALUES(?,?,?,?,?)", (tid, name, price, date, es_flash))
    audit(db, s["admin"]["username"], "precio",
          f"Creó {'VENTA FLASH' if es_flash else 'fase'} '{name}' de {t['name']}: "
          f"${price/100:.2f} desde {date}")
    db.commit()
    return jsonify(ok=True)

@app.delete("/api/admin/phases/<int:pid>")
def delete_phase(pid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    p = db.execute("SELECT p.*, t.name AS tname FROM price_phases p "
                   "JOIN ticket_types t ON t.id=p.type_id WHERE p.id=?", (pid,)).fetchone()
    if not p:
        return jsonify(error="no existe"), 404
    db.execute("DELETE FROM price_phases WHERE id=?", (pid,))
    audit(db, s["admin"]["username"], "precio",
          f"Eliminó fase '{p['name']}' de {p['tname']}")
    db.commit()
    return jsonify(ok=True)

# ------------------------------------------------ el interruptor de la venta flash

def _fase_hoy(db, tid):
    """La fase que está corriendo hoy para un tipo, saltándose las flash del
    calendario: es la que manda cuando se prende el botón."""
    hoy = now_dt().strftime("%Y-%m-%d")
    return db.execute("""SELECT * FROM price_phases WHERE type_id=? AND starts_on<=?
                         AND es_flash=0 ORDER BY starts_on DESC, id DESC LIMIT 1""",
                      (tid, hoy)).fetchone()


def estado_flash(db):
    """Lo que hay que ver ANTES de prender: en qué fase está cada tipo, a cuánto se
    vende hoy y a cuánto quedaría con el flash de esa misma fase."""
    filas = []
    for t in db.execute("SELECT * FROM ticket_types WHERE active=1 ORDER BY id").fetchall():
        fase = _fase_hoy(db, t["id"])
        normal = fase["price_cents"] if fase else t["price_cents"]
        # con fase corriendo el precio de flash es el de ESA fase; sin fase, el del
        # tipo, que es el que se está cobrando
        flash = fase["flash_price_cents"] if fase else t["flash_price_cents"]
        filas.append({
            "type_id": t["id"], "type_name": t["name"], "is_vip": bool(t["is_vip"]),
            "phase_id": fase["id"] if fase else None,
            "phase_name": fase["name"] if fase else None,
            "sin_fase": fase is None,
            "normal": money(normal),
            "flash": money(flash) if flash else None,
            "ahorro": money(normal - flash) if flash and flash < normal else 0,
            # un tipo sin precio de flash no baja al prender: se queda como está
            "listo": bool(flash and flash < normal),
        })
    return {"activa": flash_manual(db), "filas": filas}


@app.get("/api/admin/flash")
def ver_flash():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    return jsonify(**estado_flash(get_db()))


@app.post("/api/admin/flash")
def togglear_flash():
    """Prende o apaga la venta flash AHORA. No toca el calendario ni un boleto ya
    vendido: los que se generaron en flash quedaron con su precio congelado."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    activa = 1 if (request.json or {}).get("activa") else 0
    if activa:
        est = estado_flash(db)
        if not any(f["listo"] for f in est["filas"]):
            return jsonify(error="Ningún tipo tiene precio de venta flash en su fase de hoy. "
                                 "Escríbelos primero y vuelve a prender."), 400
    set_setting(db, "flash_manual", "1" if activa else "0")
    audit(db, s["admin"]["username"], "precio",
          "PRENDIÓ la venta flash" if activa else "APAGÓ la venta flash")
    db.commit()
    return jsonify(**estado_flash(db))


@app.put("/api/admin/flash")
def precios_flash():
    """El precio de flash de la fase que corre hoy, por tipo. Se guarda EN LA FASE,
    así que sigue ahí la próxima vez que se prenda dentro de esa misma fase, y cada
    fase conserva el suyo."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    precios = (request.json or {}).get("precios") or {}
    cambios = []
    for k, raw in precios.items():
        try:
            tid = int(k)
        except (TypeError, ValueError):
            continue
        t = db.execute("SELECT * FROM ticket_types WHERE id=?", (tid,)).fetchone()
        if not t:
            continue
        fase = _fase_hoy(db, tid)
        # se guarda en la fase que corre; si no hay ninguna, en el tipo
        if fase:
            sql = "UPDATE price_phases SET flash_price_cents=? WHERE id=?"
            destino, tope, donde = fase["id"], fase["price_cents"], fase["name"]
        else:
            sql = "UPDATE ticket_types SET flash_price_cents=? WHERE id=?"
            destino, tope, donde = t["id"], t["price_cents"], "precio base"
        if raw is None or str(raw).strip() == "":
            db.execute(sql, (None, destino))
            cambios.append(f"{t['name']} sin flash")
            continue
        try:
            cents = int(round(float(raw) * 100))
        except (TypeError, ValueError):
            return jsonify(error=f"Pon un precio válido para {t['name']}"), 400
        if cents <= 0:
            return jsonify(error=f"El precio de flash de {t['name']} tiene que ser mayor a cero"), 400
        if cents >= tope:
            return jsonify(error=f"El flash de {t['name']} (${cents/100:g}) no puede costar igual "
                                 f"ni más que su precio de hoy (${tope/100:g})"), 400
        db.execute(sql, (cents, destino))
        cambios.append(f"{t['name']} en flash ${cents/100:.2f} ({donde})")
    if cambios:
        audit(db, s["admin"]["username"], "precio", "Precios de venta flash: " + ", ".join(cambios))
        db.commit()
    return jsonify(**estado_flash(db))


@app.post("/api/admin/phases-all")
def create_phase_all():
    """Crea una fase para TODOS los tipos activos a la vez: misma fecha y nombre,
    con precio propio por tipo. Al llegar la fecha, todos los boletos suben de precio."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    b = request.json or {}
    name = str(b.get("name", "")).strip()
    date = str(b.get("starts_on", "")).strip()
    prices = b.get("prices") or {}
    es_flash = 1 if b.get("es_flash") else 0
    if not name or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
        return jsonify(error="Falta el nombre o la fecha (AAAA-MM-DD) de la fase"), 400
    db = get_db()
    types = db.execute("SELECT * FROM ticket_types WHERE active=1 ORDER BY id").fetchall()
    if not types:
        return jsonify(error="No hay tipos de boleto activos"), 400
    # Un tipo que se deja EN BLANCO simplemente no entra en la fase: se queda en su
    # precio base. Hace falta para los tipos que existen pero no se venden todavía
    # (Ultra VIP), que si no obligaban a inventarles un precio en cada fase.
    parsed = []
    for t in types:
        raw = prices.get(str(t["id"]), prices.get(t["id"]))
        if raw is None or str(raw).strip() == "":
            continue
        try:
            cents = int(round(float(raw) * 100))
        except (TypeError, ValueError):
            cents = 0
        if cents <= 0:
            return jsonify(error=f"Pon un precio válido para {t['name']}"), 400
        parsed.append((t["id"], cents))
    if not parsed:
        return jsonify(error="Pon al menos un precio para la fase"), 400
    for tid, cents in parsed:
        db.execute("INSERT INTO price_phases(type_id, name, price_cents, starts_on, es_flash) "
                   "VALUES(?,?,?,?,?)", (tid, name, cents, date, es_flash))
    audit(db, s["admin"]["username"], "precio",
          f"Creó la {'VENTA FLASH' if es_flash else 'fase'} '{name}' desde {date} (todos los tipos)")
    db.commit()
    return jsonify(ok=True, avisos=avisos_flash(db, parsed, date, types) if es_flash else [])

@app.put("/api/admin/phases-all")
def edit_phase_all():
    """Cambiar una fase que ya existe: su nombre, su fecha o cualquiera de sus precios.

    Antes solo se podía borrar y volver a escribirla entera. Eso convertía "quiero
    ponerle precio al Ultra VIP en la Fase 3" en un borrado seguido de teclear cuatro
    precios de memoria — y bastaba con cerrar la ventana en medio para quedarse sin la
    fase. Las fases son configuración pura: el boleto ya generado congeló su precio, así
    que reescribirlas no toca ni un peso de lo vendido."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    b = request.json or {}
    orig_name = str(b.get("orig_name", "")).strip()
    orig_date = str(b.get("orig_starts_on", "")).strip()
    name = str(b.get("name", "")).strip()
    date = str(b.get("starts_on", "")).strip()
    prices = b.get("prices") or {}
    es_flash = 1 if b.get("es_flash") else 0
    if not orig_name or not orig_date:
        return jsonify(error="Falta identificar la fase"), 400
    if not name or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
        return jsonify(error="Falta el nombre o la fecha (AAAA-MM-DD) de la fase"), 400
    db = get_db()
    if not db.execute("SELECT 1 FROM price_phases WHERE name=? AND starts_on=?",
                      (orig_name, orig_date)).fetchone():
        return jsonify(error="Esa fase ya no existe"), 404
    # si le cambian nombre y fecha, que no aterrice encima de otra fase existente
    if (name, date) != (orig_name, orig_date) and db.execute(
            "SELECT 1 FROM price_phases WHERE name=? AND starts_on=?", (name, date)).fetchone():
        return jsonify(error=f"Ya existe una fase '{name}' que arranca el {date}"), 400
    types = db.execute("SELECT * FROM ticket_types WHERE active=1 ORDER BY id").fetchall()
    parsed = []
    for t in types:
        raw = prices.get(str(t["id"]), prices.get(t["id"]))
        if raw is None or str(raw).strip() == "":
            continue      # en blanco = ese tipo no entra en la fase (se queda en su base)
        try:
            cents = int(round(float(raw) * 100))
        except (TypeError, ValueError):
            cents = 0
        if cents <= 0:
            return jsonify(error=f"Pon un precio válido para {t['name']}"), 400
        parsed.append((t["id"], cents))
    if not parsed:
        return jsonify(error="Pon al menos un precio para la fase"), 400
    antes = {r["type_id"]: r["price_cents"] for r in db.execute(
        "SELECT type_id, price_cents FROM price_phases WHERE name=? AND starts_on=?",
        (orig_name, orig_date)).fetchall()}
    db.execute("DELETE FROM price_phases WHERE name=? AND starts_on=?", (orig_name, orig_date))
    for tid, cents in parsed:
        db.execute("INSERT INTO price_phases(type_id, name, price_cents, starts_on, es_flash) "
                   "VALUES(?,?,?,?,?)", (tid, name, cents, date, es_flash))
    # el movimiento dice QUÉ cambió, no solo que se editó: si mañana un precio no
    # cuadra, la única forma de reconstruirlo es que quede escrito aquí
    nombres = {t["id"]: t["name"] for t in types}
    cambios = []
    if name != orig_name:
        cambios.append(f"nombre '{orig_name}' → '{name}'")
    if date != orig_date:
        cambios.append(f"fecha {orig_date} → {date}")
    for tid, cents in parsed:
        viejo = antes.get(tid)
        if viejo is None:
            cambios.append(f"{nombres.get(tid, tid)} ${cents/100:,.2f} (nuevo)")
        elif viejo != cents:
            cambios.append(f"{nombres.get(tid, tid)} ${viejo/100:,.2f} → ${cents/100:,.2f}")
    for tid in antes:
        if tid not in [x[0] for x in parsed]:
            cambios.append(f"{nombres.get(tid, tid)} se quitó de la fase")
    audit(db, s["admin"]["username"], "precio",
          f"Editó la fase '{orig_name}' ({orig_date}): "
          + (", ".join(cambios) if cambios else "sin cambios"))
    db.commit()
    return jsonify(ok=True, avisos=avisos_flash(db, parsed, date, types) if es_flash else [])

@app.delete("/api/admin/phases-all")
def delete_phase_all():
    """Borra una fase global (todas las filas con ese nombre y fecha)."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    name = (request.args.get("name") or "").strip()
    date = (request.args.get("starts_on") or "").strip()
    if not name or not date:
        return jsonify(error="Falta identificar la fase"), 400
    db = get_db()
    db.execute("DELETE FROM price_phases WHERE name=? AND starts_on=?", (name, date))
    audit(db, s["admin"]["username"], "precio", f"Eliminó la fase '{name}' ({date})")
    db.commit()
    return jsonify(ok=True)

@app.post("/api/admin/ticket-types")
def create_type():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    name = str(b.get("name", "")).strip()
    # Un campo vacío llega como null y float(None) revienta con 500: el panel entonces
    # dice "Error de conexión", que hace pensar en el internet cuando lo único que
    # pasó es que faltaba escribir el precio.
    try:
        price = int(round(float(b.get("price") or 0) * 100))
    except (TypeError, ValueError):
        price = 0
    if not name:
        return jsonify(error="Escribe el nombre del tipo de boleto"), 400
    if price <= 0:
        return jsonify(error="Escribe un precio mayor a cero"), 400
    # La facultad es la EXCEPCIÓN, no la regla: solo la llevan los boletos UADY. Antes
    # el valor por omisión era "sí", así que cualquier tipo nuevo nacía pidiéndola y
    # sacándola impresa en el boleto aunque nadie lo hubiera querido.
    needs_fac = 1 if b.get("needs_faculty") else 0
    # Dos tipos con el mismo nombre son un desastre silencioso: el vendedor ve dos
    # "Ultra vip" idénticos en su boletera, cada fase pide el precio dos veces, y las
    # cuentas se parten entre dos filas que parecen una. Se prohíbe repetir.
    igual = db.execute("SELECT name FROM ticket_types WHERE LOWER(TRIM(name))=LOWER(TRIM(?))",
                       (name,)).fetchone()
    if igual:
        return jsonify(error=f"Ya existe un tipo llamado «{igual['name']}». "
                             f"Edítalo en vez de crear otro."), 400
    db.execute("INSERT INTO ticket_types(name, price_cents, is_vip, needs_faculty) VALUES(?,?,?,?)",
               (name, price, 1 if b.get("is_vip") else 0, needs_fac))
    audit(db, s["admin"]["username"], "precio", f"Creó tipo '{name}' a ${price/100:.2f}")
    db.commit()
    return jsonify(ok=True)

@app.put("/api/admin/ticket-types/<int:tid>")
def edit_type(tid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    t = db.execute("SELECT * FROM ticket_types WHERE id=?", (tid,)).fetchone()
    if not t:
        return jsonify(error="no existe"), 404
    name = str(b.get("name", t["name"])).strip() or t["name"]
    if name.strip().lower() != (t["name"] or "").strip().lower():
        choca = db.execute("SELECT name FROM ticket_types WHERE LOWER(TRIM(name))=LOWER(TRIM(?)) "
                           "AND id!=?", (name, tid)).fetchone()
        if choca:
            return jsonify(error=f"Ya existe un tipo llamado «{choca['name']}»"), 400
    price = int(round(float(b.get("price", t["price_cents"] / 100)) * 100))
    active = 1 if b.get("active", t["active"]) else 0
    is_vip = 1 if b.get("is_vip", t["is_vip"]) else 0
    needs_fac = 1 if b.get("needs_faculty", t["needs_faculty"]) else 0
    db.execute("UPDATE ticket_types SET name=?, price_cents=?, active=?, is_vip=?, needs_faculty=? WHERE id=?",
               (name, price, active, is_vip, needs_fac, tid))
    if price != t["price_cents"]:
        # RF-38/90: cambio de precio auditado; boletos previos no cambian (RF-40)
        audit(db, s["admin"]["username"], "precio",
              f"Cambió precio de '{name}': ${t['price_cents']/100:.2f} → ${price/100:.2f}")
    if active != t["active"]:
        audit(db, s["admin"]["username"], "catalogo",
              f"{'Activó' if active else 'Desactivó'} tipo '{name}'")
    if needs_fac != t["needs_faculty"]:
        # cambia lo que se le pide al comprador y lo que sale impreso en el boleto
        audit(db, s["admin"]["username"], "catalogo",
              f"'{name}' {'ahora pide' if needs_fac else 'ya no pide'} facultad")
        # Decisión a mano: se apunta para que la corrección automática del arranque
        # (que quita la facultad a todo lo que no sea UADY) no la deshaga.
        manual = set(json.loads(setting(db, "facultad_manual") or "[]"))
        manual.add(tid) if needs_fac else manual.discard(tid)
        set_setting(db, "facultad_manual", json.dumps(sorted(manual)))
    db.commit()
    return jsonify(ok=True)

@app.delete("/api/admin/ticket-types/<int:tid>")
def delete_type(tid):
    """Borra un tipo de boleto y sus fases.

    Los boletos ya generados NO se tocan: llevan el nombre, el precio y la marca de
    VIP copiados en su propia fila desde el momento de la venta, así que el historial,
    las cuentas y el escáner siguen igual aunque el tipo desaparezca del catálogo.

    Lo que sí se impide es quedarse sin ningún tipo activo: los vendedores se
    quedarían con una boletera vacía."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    t = db.execute("SELECT * FROM ticket_types WHERE id=?", (tid,)).fetchone()
    if not t:
        return jsonify(error="no existe"), 404
    quedan = db.execute("SELECT COUNT(*) c FROM ticket_types WHERE id!=? AND active=1",
                        (tid,)).fetchone()["c"]
    if quedan == 0:
        return jsonify(error="Es el único tipo a la venta: si lo borras los "
                             "vendedores se quedan sin nada que vender."), 400
    vendidos = db.execute("SELECT COUNT(*) c FROM tickets WHERE type_id=? AND status!='void'",
                          (tid,)).fetchone()["c"]
    db.execute("DELETE FROM price_phases WHERE type_id=?", (tid,))
    db.execute("DELETE FROM ticket_types WHERE id=?", (tid,))
    audit(db, s["admin"]["username"], "catalogo",
          f"Eliminó el tipo de boleto '{t['name']}'" +
          (f" ({vendidos} boletos ya vendidos se conservan)" if vendidos else ""))
    db.commit()
    sync_excel_async()
    return jsonify(ok=True, sold=vendidos)

@app.get("/api/admin/faculties")
def list_faculties():
    s = require_panel()   # LEER el catálogo también el colíder: sin esto,
    # la pestaña Boletos truena al cargar sus filtros y el 401 lo saca
    # del panel con un "tu sesión terminó" que no explica nada.
    # Cambiarlo (PUT/POST/DELETE) sigue siendo solo de admins.
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    return jsonify(faculties=[dict(r) for r in
                              db.execute("SELECT * FROM faculties ORDER BY name").fetchall()])

@app.post("/api/admin/faculties")
def create_faculty():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    name = str((request.json or {}).get("name", "")).strip()
    if not name:
        return jsonify(error="Nombre requerido"), 400
    db.execute("INSERT INTO faculties(name) VALUES(?)", (name,))
    audit(db, s["admin"]["username"], "catalogo", f"Creó facultad '{name}'")
    db.commit()
    return jsonify(ok=True)

@app.put("/api/admin/faculties/<int:fid>")
def edit_faculty(fid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    f = db.execute("SELECT * FROM faculties WHERE id=?", (fid,)).fetchone()
    if not f:
        return jsonify(error="no existe"), 404
    name = str(b.get("name", f["name"])).strip() or f["name"]
    active = 1 if b.get("active", f["active"]) else 0
    db.execute("UPDATE faculties SET name=?, active=? WHERE id=?", (name, active, fid))
    audit(db, s["admin"]["username"], "catalogo", f"Editó facultad '{name}'")
    db.commit()
    return jsonify(ok=True)

# ---- gestión de vendedores (RF-82..88)

@app.get("/api/admin/sellers")
def list_sellers():
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    rows = db.execute("""
        SELECT s.*, COALESCE(SUM(CASE WHEN t.status!='void' THEN 1 ELSE 0 END),0) AS tickets,
               COUNT(t.id) AS tickets_all,
               COALESCE(SUM(CASE WHEN t.status!='void' THEN t.price_cents ELSE 0 END),0) AS total_cents
        FROM sellers s LEFT JOIN tickets t ON t.seller_id=s.id
        WHERE s.hidden=0
        GROUP BY s.id ORDER BY s.deleted, s.id""").fetchall()
    duenio = mi_ambito(s)
    if duenio:
        # el colíder solo ve a los suyos (y a sí mismo)
        rows = [r for r in rows if r["owner_admin_id"] == duenio]
    out = []
    for r in rows:
        d = dict(r)
        d["total"] = money(d.pop("total_cents"))
        d["paid"] = money(d.get("paid_cents") or 0)
        d.pop("paid_cents", None)
        d["settled"] = d["total"] > 0 and d["paid"] >= d["total"]   # Completado
        # El SERVIDOR decide quién puede tocar a este vendedor, misma verdad que en
        # las rutas: si no, el panel esconde botones que sí funcionan, o al revés.
        # Se mide con puede_cobrar porque el botón que gobierna es "Cuenta", la
        # acción del día a día — y así el colíder no ve el suyo sobre su propia ficha.
        d["can_manage"] = not r["deleted"] and puede_cobrar(db, s, r)[0]
        # el código de 4 dígitos es la credencial de acceso del vendedor: solo su
        # admin dueño lo ve (el resto sigue viendo nombre/ventas/pagos para
        # transparencia, tal como antes — solo se oculta el código)
        # El admin principal sí ve los códigos de los vendedores de un colíder: el
        # colíder trabaja PARA él, no al lado de él. Entre admins pares se siguen
        # tapando, que es para lo que se hizo esta regla.
        if not owns_seller(s["admin"], r) and not (
                es_admin_principal(s["admin"]) or duenio_es_colider(db, r)):
            d["code"] = None
        out.append(d)
    return jsonify(sellers=out)

def comision_general(db):
    """El porcentaje que se aplica a quien no tenga uno propio."""
    try:
        return max(0.0, min(100.0, float(setting(db, "seller_commission_pct") or 10)))
    except (TypeError, ValueError):
        return 10.0

def comision_pct(db, sid=None):
    """La comisión de ESTE vendedor.

    No todos llevan lo mismo: a unos se les da 10%, a otros 15 y a otros nada. El
    valor propio manda; si no tiene, se usa el general. Un 0 explícito es "sin
    comisión" y NO cae al general —por eso se compara contra None y no por verdadero
    o falso, que trataría el 0 como "sin definir"—."""
    if sid is not None:
        r = db.execute("SELECT commission_pct FROM sellers WHERE id=?", (sid,)).fetchone()
        if r is not None and r["commission_pct"] is not None:
            return max(0.0, min(100.0, float(r["commission_pct"])))
    return comision_general(db)

def vendido_cents(db, sid):
    return db.execute("""SELECT COALESCE(SUM(CASE WHEN status!='void' THEN price_cents ELSE 0 END),0) AS c
                         FROM tickets WHERE seller_id=?""", (sid,)).fetchone()["c"]

def boletos_de(db, sid):
    """Cuántos boletos lleva vendidos, sin contar los anulados. El monto solo no le
    dice nada al vendedor: "$2,575" no se compara con nada, "17 boletos" sí — es lo
    que él sabe que hizo y con lo que reclama si no le cuadra."""
    return db.execute("SELECT COUNT(*) AS n FROM tickets WHERE seller_id=? AND status!='void'",
                      (sid,)).fetchone()["n"]

def pagos_de(db, sid):
    return db.execute("SELECT * FROM seller_payments WHERE seller_id=? ORDER BY id",
                      (sid,)).fetchall()

def _pago_publico(p, saldo_despues, n):
    return {"id": p["id"], "n": n, "amount": money(p["amount_cents"]),
            "commission": money(p["commission_cents"]), "cash": money(p["cash_cents"]),
            "commission_pct": p["commission_pct"], "note": p["note"] or "",
            "created_by": p["created_by"], "created_at": p["created_at"],
            "balance_after": money(saldo_despues)}

def estado_cuenta(db, sid):
    """Cuenta completa del vendedor: cuánto vendió, cuánto ha abonado, cuánto se ha
    llevado de comisión y el saldo tras CADA pago (para poder mostrarle el recorrido)."""
    vendido = vendido_cents(db, sid)
    filas = pagos_de(db, sid)
    abonado = com = efectivo = 0
    historial = []
    for i, p in enumerate(filas, 1):
        abonado += p["amount_cents"]
        com += p["commission_cents"]
        efectivo += p["cash_cents"]
        # el número es el orden real del corte (1 = el primero que pagó), por eso se
        # calcula sobre la lista vieja→nueva y no cambia aunque después se invierta
        historial.append(_pago_publico(p, vendido - abonado, i))
    historial.reverse()          # el más reciente arriba
    return {"sold": money(vendido), "sold_tickets": boletos_de(db, sid),
            "settled_amount": money(abonado),
            "commission_total": money(com), "cash_total": money(efectivo),
            "balance": money(vendido - abonado),
            "settled": vendido > 0 and abonado >= vendido,
            "commission_pct": comision_pct(db, sid),
            "commission_general": comision_general(db),
            "commission_propia": (lambda r: r and r["commission_pct"] is not None)(
                db.execute("SELECT commission_pct FROM sellers WHERE id=?", (sid,)).fetchone()),
            "payments": historial}

@app.get("/api/admin/sellers/<int:sid>/payments")
def list_seller_payments(sid):
    """Historial de pagos de un vendedor. Lo ven TODOS los admins (transparencia),
    igual que sus ventas; solo su admin dueño puede registrar o borrar."""
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    sel = db.execute("SELECT * FROM sellers WHERE id=? AND hidden=0", (sid,)).fetchone()
    if not sel:
        return jsonify(error="no existe"), 404
    # un colíder no puede ni ASOMARSE a la cuenta de un vendedor que no es suyo
    if es_colider(s) and sel["owner_admin_id"] != s["admin"]["id"]:
        return jsonify(error="no existe"), 404
    out = estado_cuenta(db, sid)
    out["seller_name"] = sel["name"]
    # ver su propia cuenta sí (tiene que saber cuánto debe); cobrársela él, no
    out["can_edit"] = puede_cobrar(db, s, sel)[0]
    # La comisión la fija un admin, punto. El colíder cobra a su equipo pero no
    # decide cuánto gana nadie —ni ellos ni él—. Va aparte de can_edit porque él SÍ
    # cobra: si se mezclaran, o se le quita el cobro o se le regala el sueldo.
    out["can_commission"] = (not es_colider(s)) and puede_gestionar(db, s["admin"], sel)
    return jsonify(**out)

@app.post("/api/admin/sellers/<int:sid>/payments")
def add_seller_payment(sid):
    """Registra UNA entrega de dinero del vendedor.

    El admin captura el ABONO (cuánto de la deuda se salda). De ahí sale sola la
    comisión que el vendedor se queda y el efectivo que debe entregar:
        abono $20,000 · comisión 10% = $2,000 · efectivo a recibir $18,000
    Cada pago queda con su fecha, para poder demostrarle después cómo fue pagando."""
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # hidden=0: el vendedor de invitados no se toca desde el panel
    sel = db.execute("SELECT * FROM sellers WHERE id=? AND deleted=0 AND hidden=0",
                     (sid,)).fetchone()
    if not sel:
        return jsonify(error="no existe"), 404
    ok, motivo = puede_cobrar(db, s, sel)
    if not ok:
        return jsonify(error=motivo), 403
    b = request.json or {}
    try:
        abono = int(round(float(b.get("amount", 0)) * 100))
    except (TypeError, ValueError):
        return jsonify(error="Monto inválido"), 400
    if abono <= 0:
        return jsonify(error="El abono debe ser mayor a cero"), 400
    vendido = vendido_cents(db, sid)
    ya = sum(p["amount_cents"] for p in pagos_de(db, sid))
    if vendido <= 0:
        return jsonify(error="Este vendedor aún no ha vendido nada; no hay pago que registrar"), 400
    if ya + abono > vendido:
        falta = (vendido - ya) / 100
        return jsonify(error=f"Se pasa de lo que debe. Su saldo pendiente es ${falta:,.2f}"), 400
    pct = comision_pct(db, sid)
    comision = int(round(abono * pct / 100))
    efectivo = abono - comision
    db.execute("""INSERT INTO seller_payments
        (seller_id, seller_name, amount_cents, commission_cents, cash_cents,
         commission_pct, note, created_by, created_at)
        VALUES(?,?,?,?,?,?,?,?,?)""",
        (sid, sel["name"], abono, comision, efectivo, pct,
         str(b.get("note", "")).strip()[:120] or None, s["admin"]["username"], now_iso()))
    # paid_cents queda como espejo del total abonado, para que el resumen y la
    # lista de vendedores (que ya lo usaban) sigan cuadrando sin cambios
    db.execute("UPDATE sellers SET paid_cents=? WHERE id=?", (ya + abono, sid))
    saldo = vendido - (ya + abono)
    audit(db, s["admin"]["username"], "pago",
          f"Recibió ${efectivo/100:,.2f} de '{sel['name']}' (abono ${abono/100:,.2f}, "
          f"comisión ${comision/100:,.2f}) · saldo ${saldo/100:,.2f}")
    db.commit()
    out = estado_cuenta(db, sid)
    out["seller_name"] = sel["name"]
    out["can_edit"] = True
    return jsonify(**out)

@app.get("/api/admin/sellers/<int:sid>/payments.xlsx")
def export_seller_payments(sid):
    """Estado de cuenta del vendedor en Excel, para llevar el control o mandárselo
    si pide cuentas. Mismo detalle que la imagen, pero en hoja de cálculo."""
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    sel = db.execute("SELECT * FROM sellers WHERE id=? AND hidden=0", (sid,)).fetchone()
    if not sel:
        return jsonify(error="no existe"), 404
    # el mismo candado que en el estado de cuenta en pantalla: si no se pusiera aquí,
    # bastaba con pedir el Excel por número para leer las cuentas de otro grupo
    if es_colider(s) and sel["owner_admin_id"] != s["admin"]["id"]:
        return jsonify(error="no existe"), 404
    c = estado_cuenta(db, sid)
    comision_total = c["sold"] * c["commission_pct"] / 100
    debe_entregar = c["sold"] - comision_total

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de cuenta"
    titulo = Font(bold=True, size=14)
    etiqueta = Font(bold=True)
    dinero = '"$"#,##0.00'

    ws["A1"] = f"Estado de cuenta · {sel['name']}"; ws["A1"].font = titulo
    ws["A2"] = "Generado el " + now_dt().strftime("%d/%m/%Y %H:%M")
    ws.cell(row=3, column=1, value="Boletos vendidos").font = etiqueta
    ws.cell(row=3, column=2, value=c["sold_tickets"])
    resumen = [
        ("Vendió en boletos", c["sold"]),
        (f"Su comisión ({c['commission_pct']:g}%)", -comision_total),
        ("Debe entregar en total", debe_entregar),
        ("Ya entregó", c["cash_total"]),
        ("Le falta entregar", max(0, debe_entregar - c["cash_total"])),
    ]
    fila = 4
    for etq, val in resumen:
        ws.cell(row=fila, column=1, value=etq).font = etiqueta
        celda = ws.cell(row=fila, column=2, value=val)
        celda.number_format = dinero
        fila += 1

    fila += 1
    encabezados = ["Corte", "Fecha", "Efectivo entregado", "Cubrió de su cuenta",
                   "Su comisión", "Quedó debiendo", "Nota", "Registró"]
    for col, h in enumerate(encabezados, 1):
        cel = ws.cell(row=fila, column=col, value=h)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="C0501E")
    # del más viejo al más nuevo: se lee como fue pagando
    for p in reversed(c["payments"]):
        fila += 1
        ws.cell(row=fila, column=1, value=f"Pago {p['n']}").font = etiqueta
        ws.cell(row=fila, column=2, value=p["created_at"])
        for col, val in enumerate([p["cash"], p["amount"], p["commission"],
                                   p["balance_after"]], 3):
            cel = ws.cell(row=fila, column=col, value=val)
            cel.number_format = dinero
        ws.cell(row=fila, column=7, value=p["note"])
        ws.cell(row=fila, column=8, value=p["created_by"])
    for col, ancho in enumerate([10, 19, 19, 20, 14, 17, 26, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = re.sub(r"[^\w]+", "_", sel["name"]).strip("_")[:40] or "vendedor"
    return send_file(buf, as_attachment=True, download_name=f"cuenta_{slug}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.delete("/api/admin/payments/<int:pid>")
def delete_seller_payment(pid):
    """Borra un pago mal capturado. Queda registrado en Movimientos: el historial
    debe poder corregirse, pero nunca en silencio."""
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    p = db.execute("SELECT * FROM seller_payments WHERE id=?", (pid,)).fetchone()
    if not p:
        return jsonify(error="no existe"), 404
    sel = db.execute("SELECT * FROM sellers WHERE id=?", (p["seller_id"],)).fetchone()
    if sel is not None:
        ok, motivo = puede_cobrar(db, s, sel)
        if not ok:
            return jsonify(error=motivo.replace("registrar pagos de", "borrar pagos de")), 403
    db.execute("DELETE FROM seller_payments WHERE id=?", (pid,))
    total = sum(x["amount_cents"] for x in pagos_de(db, p["seller_id"]))
    db.execute("UPDATE sellers SET paid_cents=? WHERE id=?", (total, p["seller_id"]))
    audit(db, s["admin"]["username"], "pago",
          f"Borró un pago de '{p['seller_name']}': abono ${p['amount_cents']/100:,.2f} "
          f"del {p['created_at'][:10]}")
    db.commit()
    return jsonify(ok=True)

@app.post("/api/admin/sellers/bulk")
def create_sellers_bulk():
    """Alta de varios vendedores de un jalón, pegando la lista de nombres.

    Con 50 vendedores, capturarlos uno por uno son 50 formularios y 50 códigos
    copiados a mano. Aquí se pegan los nombres —uno por línea— y salen todos con su
    código listo para repartir. Los repetidos NO frenan la carga: se apartan y se
    reportan, para que 49 altas buenas no se pierdan por un nombre duplicado."""
    # El colíder también: pegar 20 nombres es lo mismo que teclear 20 altas, y si
    # puede lo segundo negarle lo primero no protege nada — solo lo saca del panel,
    # porque cualquier 401 aquí se lee como "tu sesión terminó".
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    crudo = str((request.json or {}).get("names", ""))
    nombres, vistos = [], set()
    for linea in crudo.splitlines():
        n = re.sub(r"\s+", " ", linea).strip()
        if not n:
            continue
        if n.lower() in vistos:          # repetido DENTRO de la misma lista
            continue
        vistos.add(n.lower())
        nombres.append(n)
    if not nombres:
        return jsonify(error="Pega al menos un nombre"), 400
    if len(nombres) > 200:
        return jsonify(error="Máximo 200 por vez"), 400

    creados, repetidos = [], []
    for n in nombres:
        ya = db.execute("SELECT 1 FROM sellers WHERE deleted=0 AND hidden=0 "
                        "AND LOWER(TRIM(name))=LOWER(TRIM(?))", (n,)).fetchone()
        if ya:
            repetidos.append(n)
            continue
        code = gen_seller_code(db)
        db.execute("INSERT INTO sellers(name, code, owner_admin_id, owner_admin_name, created_at) "
                   "VALUES(?,?,?,?,?)",
                   (n, code, s["admin"]["id"], s["admin"]["username"], now_iso()))
        creados.append({"name": n, "code": code})
    if creados:
        audit(db, s["admin"]["username"], "vendedor_creado",
              f"Alta masiva: {len(creados)} vendedores")
    db.commit()
    return jsonify(ok=True, creados=creados, repetidos=repetidos)

@app.post("/api/admin/sellers")
def create_seller():
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    # se normalizan los espacios: "Luis  Pérez " y "Luis Pérez" son la misma persona,
    # y de paso el nombre no sale con huecos raros en el boleto
    name = re.sub(r"\s+", " ", str(b.get("name", ""))).strip()
    if not name:
        return jsonify(error="Nombre requerido"), 400
    code = str(b.get("code", "")).strip()
    if code:
        if not re.fullmatch(r"\d{5}", code):
            return jsonify(error="El código debe ser de 5 dígitos"), 400
        if db.execute("SELECT 1 FROM sellers WHERE code=? AND deleted=0", (code,)).fetchone():
            return jsonify(error="Ese código ya está en uso"), 400   # RF-84
    else:
        code = gen_seller_code(db)
    # Dos vendedores con el mismo nombre son una trampa al cobrar: se abre la cuenta
    # equivocada y parece que no ha vendido nada. No se prohíbe (puede haber dos
    # Luis de verdad), pero hay que confirmarlo a propósito.
    if not b.get("force"):
        # solo se compara contra los vendedores de quien está dando de alta: avisarle
        # a un colíder que "ya existe un Luis" cuando el Luis es de otro grupo le
        # confirma vendedores que no tiene por qué conocer
        ambito = mi_ambito(s)
        igual = db.execute(
            "SELECT code FROM sellers WHERE deleted=0 AND hidden=0 "
            "AND LOWER(TRIM(name))=LOWER(TRIM(?))"
            + (" AND owner_admin_id=?" if ambito else ""),
            ((name, ambito) if ambito else (name,))).fetchone()
        if igual:
            return jsonify(error=f"Ya tienes un vendedor llamado «{name}» "
                                 f"(código {igual['code']}).", duplicate=True), 409
    # el vendedor queda ligado al admin que lo crea (su dueño)
    db.execute("INSERT INTO sellers(name, code, owner_admin_id, owner_admin_name, created_at) "
               "VALUES(?,?,?,?,?)",
               (name, code, s["admin"]["id"], s["admin"]["username"], now_iso()))
    audit(db, s["admin"]["username"], "vendedor_creado",
          f"Creó al vendedor '{name}' (código {code})")
    db.commit()
    return jsonify(ok=True, code=code)

@app.put("/api/admin/sellers/<int:sid>")
def edit_seller(sid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    # hidden=0: el vendedor de invitados no se puede tocar desde el panel (si se
    # borrara, sus boletos viejos reaparecerían como venta en el resumen)
    sel = db.execute("SELECT * FROM sellers WHERE id=? AND deleted=0 AND hidden=0",
                     (sid,)).fetchone()
    if not sel:
        return jsonify(error="no existe"), 404
    if not puede_gestionar(db, s["admin"], sel):
        return jsonify(error=f"Solo {sel['owner_admin_name']} (su admin) puede modificar a este vendedor"), 403
    name = str(b.get("name", sel["name"])).strip() or sel["name"]
    code = str(b.get("code", sel["code"])).strip()
    if code != sel["code"]:
        if not re.fullmatch(r"\d{5}", code):
            return jsonify(error="El código debe ser de 5 dígitos"), 400
        if db.execute("SELECT 1 FROM sellers WHERE code=? AND deleted=0 AND id!=?",
                      (code, sid)).fetchone():
            return jsonify(error="Ese código ya está en uso"), 400
        db.execute("DELETE FROM sessions WHERE role='seller' AND user_id=?", (sid,))
    # Solo se escribe en Movimientos si de verdad cambió algo. Ahora la comisión se
    # toca desde la cuenta del vendedor sin pasar por este formulario, y anotar
    # "editó vendedor" en cada ajuste de porcentaje llenaba el historial de ruido.
    if name != sel["name"] or code != sel["code"]:
        db.execute("UPDATE sellers SET name=?, code=? WHERE id=?", (name, code, sid))
        audit(db, s["admin"]["username"], "usuarios",
              f"Editó vendedor '{sel['name']}' → nombre '{name}', código {code}")

    # Comisión propia. Se manda "" (o null) para que vuelva a usar la general, y un
    # número —incluido el 0— para fijarla. Los pagos YA registrados no cambian:
    # cada uno guardó el porcentaje con el que se hizo.
    if "commission_pct" in b:
        crudo = b.get("commission_pct")
        if crudo is None or str(crudo).strip() == "":
            nueva = None
        else:
            try:
                nueva = max(0.0, min(100.0, float(crudo)))
            except (TypeError, ValueError):
                return jsonify(error="La comisión debe ser un número entre 0 y 100"), 400
        if nueva != sel["commission_pct"]:
            db.execute("UPDATE sellers SET commission_pct=? WHERE id=?", (nueva, sid))
            audit(db, s["admin"]["username"], "usuarios",
                  f"Comisión de '{name}': " +
                  (f"{nueva:g}%" if nueva is not None
                   else f"la general ({comision_general(db):g}%)"))
    db.commit()
    return jsonify(ok=True)

@app.post("/api/admin/sellers/<int:sid>/toggle")
def toggle_seller(sid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # hidden=0: el vendedor de invitados no se puede tocar desde el panel (si se
    # borrara, sus boletos viejos reaparecerían como venta en el resumen)
    sel = db.execute("SELECT * FROM sellers WHERE id=? AND deleted=0 AND hidden=0",
                     (sid,)).fetchone()
    if not sel:
        return jsonify(error="no existe"), 404
    if not puede_gestionar(db, s["admin"], sel):
        return jsonify(error=f"Solo {sel['owner_admin_name']} (su admin) puede modificar a este vendedor"), 403
    new = 0 if sel["active"] else 1
    db.execute("UPDATE sellers SET active=? WHERE id=?", (new, sid))
    if not new:
        # RF-32/86: cerrar sesión de inmediato
        db.execute("DELETE FROM sessions WHERE role='seller' AND user_id=?", (sid,))
    audit(db, s["admin"]["username"], "usuarios",
          f"{'Reactivó' if new else 'Desactivó'} al vendedor '{sel['name']}'")
    db.commit()
    return jsonify(ok=True, active=bool(new))

@app.delete("/api/admin/sellers/<int:sid>")
def delete_seller(sid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # hidden=0: el vendedor de invitados no se puede tocar desde el panel (si se
    # borrara, sus boletos viejos reaparecerían como venta en el resumen)
    sel = db.execute("SELECT * FROM sellers WHERE id=? AND deleted=0 AND hidden=0",
                     (sid,)).fetchone()
    if not sel:
        return jsonify(error="no existe"), 404
    if not puede_gestionar(db, s["admin"], sel):
        return jsonify(error=f"Solo {sel['owner_admin_name']} (su admin) puede eliminar a este vendedor"), 403
    n = db.execute("SELECT COUNT(*) c FROM tickets WHERE seller_id=?", (sid,)).fetchone()["c"]
    # RF-87: se elimina la cuenta, los boletos se conservan con su nombre
    db.execute("UPDATE sellers SET deleted=1, active=0, code=NULL WHERE id=?", (sid,))
    db.execute("DELETE FROM sessions WHERE role='seller' AND user_id=?", (sid,))
    audit(db, s["admin"]["username"], "vendedor_eliminado",
          f"Eliminó al vendedor '{sel['name']}' ({n} boletos quedan asociados a su nombre)")
    db.commit()
    sync_excel_async()
    return jsonify(ok=True, tickets_kept=n)

# ---- administradores (RF-89, RF-35, RF-36)

@app.get("/api/admin/admins")
def list_admins():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    rows = db.execute("SELECT id, username, created_at, role FROM admins ORDER BY id").fetchall()
    return jsonify(admins=[dict(r) for r in rows], me=s["admin"]["id"])

@app.post("/api/admin/admins")
def create_admin():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    username = str(b.get("username", "")).strip()
    password = str(b.get("password", ""))
    if len(username) < 3 or len(password) < 8:
        return jsonify(error="Usuario mín. 3 caracteres y contraseña mín. 8"), 400
    if db.execute("SELECT 1 FROM admins WHERE username=?", (username,)).fetchone():
        return jsonify(error="Ese usuario ya existe"), 400
    rol = "colider" if b.get("role") == "colider" else "admin"
    db.execute("INSERT INTO admins(username, pass_hash, created_at, role) VALUES(?,?,?,?)",
               (username, hash_password(password), now_iso(), rol))
    code = None
    reusado = False
    if rol == "colider":
        # El colíder también vende en persona, así que necesita ficha de vendedor
        # marcada es_lider: sus ventas personales se cuentan y comisionan aparte de
        # las de su equipo, sin inventar un caso especial.
        nuevo = db.execute("SELECT id FROM admins WHERE username=?", (username,)).fetchone()
        # Casi siempre el colíder sale de entre los vendedores: ya venía vendiendo y
        # ya repartió su código. Abrirle uno nuevo lo deja con dos —y con sus ventas
        # partidas en dos cuentas—, así que si dicen cuál es el suyo, se reaprovecha
        # ese: conserva su código, su historial y lo que se le debe.
        actual = str(b.get("seller_code", "")).strip()
        if actual:
            sel = db.execute("SELECT * FROM sellers WHERE code=? AND deleted=0 AND hidden=0",
                             (actual,)).fetchone()
            if not sel:
                return jsonify(error=f"No hay ningún vendedor activo con el código {actual}"), 400
            # esa ficha ya es la personal de otro colíder: reasignarla lo dejaría a él
            # con el grupo descabezado y sus ventas contadas como de otro
            if sel["es_lider"]:
                dueno = db.execute("SELECT username FROM admins WHERE id=?",
                                   (sel["owner_admin_id"],)).fetchone()
                return jsonify(error=f"El código {actual} ya es el del colíder "
                                     f"{dueno['username'] if dueno else 'otro'}."), 400
            db.execute("UPDATE sellers SET owner_admin_id=?, owner_admin_name=?, es_lider=1 "
                       "WHERE id=?", (nuevo["id"], username, sel["id"]))
            code, reusado = sel["code"], True
        else:
            code = gen_seller_code(db)
            db.execute("INSERT INTO sellers(name, code, owner_admin_id, owner_admin_name, "
                       "created_at, es_lider) VALUES(?,?,?,?,?,1)",
                       (username, code, nuevo["id"], username, now_iso()))
    audit(db, s["admin"]["username"], "usuarios",
          f"Creó {'colíder' if rol == 'colider' else 'administrador'} '{username}'"
          + (f" (su código de vendedor: {code})" if code else "")
          + (" — se le conservó el que ya tenía" if reusado else ""))
    db.commit()
    return jsonify(ok=True, code=code, reusado=reusado)

@app.delete("/api/admin/admins/<int:aid>")
def delete_admin(aid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    if aid == s["admin"]["id"]:
        return jsonify(error="No puedes eliminarte a ti mismo"), 400   # RF-35
    n = db.execute("SELECT COUNT(*) c FROM admins").fetchone()["c"]
    if n <= 1:
        return jsonify(error="No se puede borrar el último administrador"), 400   # RF-36
    target = db.execute("SELECT * FROM admins WHERE id=?", (aid,)).fetchone()
    if not target:
        return jsonify(error="no existe"), 404
    # Su equipo NO se puede quedar apuntando a un admin que ya no existe. Si eso
    # pasa, owns_seller() le dice que no a todo el mundo y esos vendedores quedan
    # trabados para siempre: no se les puede cobrar, ni editar, ni anular un boleto
    # suyo. Pasan a quien lo está borrando, que para eso es el jefe.
    # Primero su ficha personal, que se va con él pero SOLO si no vendió nada:
    # borrar una que ya tiene boletos dejaría dinero cobrado sin dueño. (Va antes de
    # reasignar, mientras es_lider todavía dice cuál es la suya.)
    db.execute("DELETE FROM sellers WHERE owner_admin_id=? AND es_lider=1 "
               "AND id NOT IN (SELECT seller_id FROM tickets WHERE seller_id IS NOT NULL)",
               (aid,))
    db.execute("UPDATE sellers SET owner_admin_id=?, owner_admin_name=?, es_lider=0 "
               "WHERE owner_admin_id=?",
               (s["admin"]["id"], s["admin"]["username"], aid))
    db.execute("DELETE FROM admins WHERE id=?", (aid,))
    db.execute("DELETE FROM sessions WHERE role='admin' AND user_id=?", (aid,))
    audit(db, s["admin"]["username"], "usuarios", f"Eliminó administrador '{target['username']}'")
    db.commit()
    return jsonify(ok=True)

@app.get("/api/admin/grupos")
def grupos_colider():
    """Cada colíder con sus números partidos en dos: lo que vendió ÉL en persona y lo
    que vendieron SUS vendedores. Sumar todo en un solo número esconde justo lo que se
    quiere saber —si el colíder trabaja o solo administra—, así que van separados y
    con el total al lado."""
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    duenio = mi_ambito(s)
    lideres = db.execute("SELECT id, username FROM admins WHERE role='colider' ORDER BY username").fetchall()
    if duenio:
        lideres = [l for l in lideres if l["id"] == duenio]
    salida = []
    for l in lideres:
        filas = db.execute(f"""
            SELECT s.id, s.name, s.code, s.es_lider, s.deleted, s.paid_cents,
                   COUNT(t.id) AS n,
                   COALESCE(SUM(CASE WHEN t.status!='void' THEN t.price_cents ELSE 0 END),0) AS cents
            FROM sellers s LEFT JOIN tickets t ON t.seller_id=s.id AND t.status!='void' 
            WHERE s.hidden=0 AND s.owner_admin_id=?
            GROUP BY s.id ORDER BY s.es_lider DESC, cents DESC, s.name ASC""",
            (l["id"],)).fetchall()
        propio = [r for r in filas if r["es_lider"]]
        equipo = [r for r in filas if not r["es_lider"]]
        def suma(rs, campo):
            return sum(r[campo] or 0 for r in rs)
        salida.append(dict(
            id=l["id"], nombre=l["username"],
            propio=dict(boletos=suma(propio, "n"), monto=money(suma(propio, "cents")),
                        cobrado=money(suma(propio, "paid_cents")),
                        code=(propio[0]["code"] if propio else None)),
            equipo=dict(boletos=suma(equipo, "n"), monto=money(suma(equipo, "cents")),
                        cobrado=money(suma(equipo, "paid_cents")),
                        vendedores=len([r for r in equipo if not r["deleted"]])),
            total=dict(boletos=suma(filas, "n"), monto=money(suma(filas, "cents")),
                       cobrado=money(suma(filas, "paid_cents"))),
            miembros=[dict(id=r["id"], name=r["name"], code=r["code"],
                           es_lider=bool(r["es_lider"]), deleted=bool(r["deleted"]),
                           boletos=r["n"] or 0, monto=money(r["cents"] or 0),
                           cobrado=money(r["paid_cents"] or 0)) for r in filas],
        ))
    return jsonify(grupos=salida, soy_colider=bool(duenio))

@app.get("/api/admin/cortesias")
def listar_cortesias():
    """Los invitados especiales, aparte de la venta.

    No entran en el resumen ni en la cobranza —no pagaron, no hay dinero que cuadrar—
    pero el día del evento hacen falta a la vista: quién ya entró, quién no llegó y a
    qué hora pasó. Eso no se puede saber desde la lista de boletos vendidos, que es
    donde antes se mezclaban."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # Las anuladas SÍ salen en la lista, tachadas: si desaparecieran, quien la anuló
    # no tendría cómo comprobar que lo hizo, y un invitado que llega a reclamar deja
    # al de la puerta sin nada que enseñarle. Lo que no hacen es contar.
    rows = db.execute("SELECT * FROM tickets WHERE es_cortesia=1 "
                      "ORDER BY status='void', used_at IS NULL DESC, used_at DESC, id DESC").fetchall()
    out = []
    for t in rows:
        d = ticket_public(t)
        d["entro"] = bool(t["used_at"]) and t["status"] != "void"
        d["anulada"] = t["status"] == "void"
        d["void_reason"] = t["void_reason"] or ""
        out.append(d)
    vivas = [x for x in out if not x["anulada"]]
    return jsonify(cortesias=out, total=len(vivas),
                   entraron=sum(1 for x in vivas if x["entro"]),
                   faltan=sum(1 for x in vivas if not x["entro"]),
                   anuladas=len(out) - len(vivas))

# ---- auditoría, exportación, ajustes

@app.get("/api/admin/audit")
def get_audit():
    s = require_panel()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    # Movimientos = solo acciones de ADMINS. La generación de boletos la hacen los
    # vendedores y ya se refleja en Ventas/Vendedores, así que no se lista aquí.
    #
    # El colíder ve SOLO lo que él mismo hizo. El registro es global —cambios de
    # precio, anulaciones, cobros de otros grupos— y enseñárselo entero sería darle
    # por la puerta de atrás justo lo que no debe ver.
    if es_colider(s):
        rows = db.execute("SELECT * FROM audit_log WHERE action != 'generacion' "
                          "AND actor = ? ORDER BY id DESC LIMIT 500",
                          (s["admin"]["username"],)).fetchall()
    else:
        rows = db.execute("SELECT * FROM audit_log WHERE action != 'generacion' "
                          "ORDER BY id DESC LIMIT 500").fetchall()
    return jsonify(log=[dict(r) for r in rows])

# ---- gastos de la fiesta (local, bebidas, DJ, etc.): cuánto se debe -------------

def sana_gastos(db):
    """Un gasto marcado 'pagado' de antes tiene que arrancar con su abono completo,
    o el día que se despliegue esto todos los pagados aparecerían debiendo todo."""
    db.execute("UPDATE expenses SET paid_cents=amount_cents "
               "WHERE status='pagado' AND paid_cents<amount_cents")
    db.execute("UPDATE expenses SET paid_cents=amount_cents WHERE paid_cents>amount_cents")

def estado_gasto(amount, paid):
    if paid >= amount and amount > 0:
        return "pagado"
    return "abonado" if paid > 0 else "pendiente"

def expenses_summary(db):
    rows = db.execute("SELECT * FROM expenses ORDER BY id DESC").fetchall()
    total = paid = pending = 0
    by_account = {}
    for r in rows:
        c = r["amount_cents"] or 0
        # lo abonado nunca puede pasarse del gasto: si alguien bajó el monto después
        # de pagar de más, se topa aquí y no se inventa un saldo a favor
        pc = min(r["paid_cents"] or 0, c)
        total += c
        paid += pc
        pending += c - pc
        acc = (r["account"] or "").strip() or "Sin asignar"
        b = by_account.setdefault(acc, {"total": 0, "paid": 0, "pending": 0})
        b["total"] += c
        b["paid"] += pc
        b["pending"] += c - pc
    accounts = [{"account": k, "total": money(v["total"]), "paid": money(v["paid"]),
                 "pending": money(v["pending"])}
                for k, v in sorted(by_account.items(), key=lambda kv: -kv[1]["total"])]
    items = [{"id": r["id"], "name": r["name"], "amount": money(r["amount_cents"]),
              "account": r["account"] or "",
              "paid": money(min(r["paid_cents"] or 0, r["amount_cents"] or 0)),
              "pending": money(max(0, (r["amount_cents"] or 0) - (r["paid_cents"] or 0))),
              "status": estado_gasto(r["amount_cents"] or 0, r["paid_cents"] or 0),
              "created_at": r["created_at"]} for r in rows]
    return {"expenses": items, "total": money(total), "paid": money(paid),
            "pending": money(pending), "by_account": accounts}

@app.get("/api/admin/expenses")
def list_expenses():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    sana_gastos(db); db.commit()
    return jsonify(**expenses_summary(db))

@app.post("/api/admin/expenses")
def create_expense():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    b = request.json or {}
    name = str(b.get("name", "")).strip()
    if not name:
        return jsonify(error="Escribe el nombre del gasto"), 400
    try:
        cents = int(round(float(b.get("amount", 0)) * 100))
    except (TypeError, ValueError):
        cents = 0
    if cents < 0:
        return jsonify(error="El monto no puede ser negativo"), 400
    account = str(b.get("account", "")).strip()
    status = "pagado" if b.get("status") == "pagado" else "pendiente"
    ya = cents if status == "pagado" else 0
    db = get_db()
    db.execute("INSERT INTO expenses(name, amount_cents, account, status, paid_cents, "
               "created_by, created_at) VALUES(?,?,?,?,?,?,?)",
               (name, cents, account, status, ya, s["admin"]["username"], now_iso()))
    audit(db, s["admin"]["username"], "gasto",
          f"Agregó gasto '{name}': ${cents/100:,.2f} ({status}{', ' + account if account else ''})")
    db.commit()
    return jsonify(ok=True)

@app.put("/api/admin/expenses/<int:eid>")
def edit_expense(eid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    e = db.execute("SELECT * FROM expenses WHERE id=?", (eid,)).fetchone()
    if not e:
        return jsonify(error="no existe"), 404
    b = request.json or {}
    name = str(b.get("name", e["name"])).strip() or e["name"]
    try:
        cents = int(round(float(b.get("amount", e["amount_cents"] / 100)) * 100))
    except (TypeError, ValueError):
        cents = e["amount_cents"]
    if cents < 0:
        cents = 0
    account = str(b.get("account", e["account"] or "")).strip()
    status = b.get("status", e["status"])
    status = "pagado" if status == "pagado" else "pendiente"
    # "Marcar pagado" salda lo que falte; quitar el pagado NO borra los abonos que
    # de verdad se entregaron —eso sería perder dinero de la cuenta—, solo deja de
    # darlo por liquidado. Y si bajan el monto, lo abonado se topa ahí.
    ya = min(e["paid_cents"] or 0, cents)
    if status == "pagado":
        ya = cents
    elif (e["paid_cents"] or 0) >= (e["amount_cents"] or 0) and e["status"] == "pagado":
        ya = 0
    db.execute("UPDATE expenses SET name=?, amount_cents=?, account=?, status=?, paid_cents=? "
               "WHERE id=?", (name, cents, account, status, ya, eid))
    if status != e["status"]:
        audit(db, s["admin"]["username"], "gasto",
              f"Marcó '{name}' como {status}")
    else:
        audit(db, s["admin"]["username"], "gasto", f"Editó gasto '{name}'")
    db.commit()
    return jsonify(ok=True)

@app.post("/api/admin/expenses/<int:eid>/abono")
def abonar_gasto(eid):
    """Un adelanto: del local de $15,000 se entregaron $3,000. Lo que sigue debiendo
    sale solo. Cada abono queda en Movimientos con quién y cuánto: es dinero que sale
    de la bolsa y tiene que poder reconstruirse después."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    e = db.execute("SELECT * FROM expenses WHERE id=?", (eid,)).fetchone()
    if not e:
        return jsonify(error="no existe"), 404
    try:
        abono = int(round(float((request.json or {}).get("amount", 0)) * 100))
    except (TypeError, ValueError):
        return jsonify(error="Monto inválido"), 400
    if abono == 0:
        return jsonify(error="El abono no puede ser cero"), 400
    ya = min(e["paid_cents"] or 0, e["amount_cents"] or 0)
    nuevo = ya + abono
    if nuevo < 0:
        return jsonify(error="No puedes quitar más de lo que se ha abonado"), 400
    if nuevo > (e["amount_cents"] or 0):
        falta = ((e["amount_cents"] or 0) - ya) / 100
        return jsonify(error=f"Se pasa del gasto. Falta ${falta:,.2f} por pagar"), 400
    estado = "pagado" if nuevo >= (e["amount_cents"] or 0) and e["amount_cents"] else "pendiente"
    db.execute("UPDATE expenses SET paid_cents=?, status=? WHERE id=?", (nuevo, estado, eid))
    audit(db, s["admin"]["username"], "gasto",
          (f"Abonó ${abono/100:,.2f} a '{e['name']}'" if abono > 0
           else f"Corrigió ${-abono/100:,.2f} de lo abonado a '{e['name']}'")
          + f" — lleva ${nuevo/100:,.2f} de ${(e['amount_cents'] or 0)/100:,.2f}")
    db.commit()
    return jsonify(ok=True, paid=money(nuevo),
                   pending=money(max(0, (e["amount_cents"] or 0) - nuevo)))

@app.delete("/api/admin/expenses/<int:eid>")
def delete_expense(eid):
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    e = db.execute("SELECT * FROM expenses WHERE id=?", (eid,)).fetchone()
    if not e:
        return jsonify(error="no existe"), 404
    db.execute("DELETE FROM expenses WHERE id=?", (eid,))
    audit(db, s["admin"]["username"], "gasto", f"Eliminó gasto '{e['name']}'")
    db.commit()
    return jsonify(ok=True)

@app.get("/api/admin/export")
def export_xlsx():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    where, params = ticket_filters()   # RF-93: respeta filtros
    rows = db.execute("SELECT * FROM tickets" + where + " ORDER BY id", params).fetchall()
    wb = build_workbook(rows, seller_summary(db))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fdesc = " con filtros" if where else " completa"
    audit(db, s["admin"]["username"], "exportacion",
          f"Exportó la base de compradores{fdesc} ({len(rows)} boletos)")   # RF-94
    db.commit()
    name = f"boletos_{now_dt().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/admin/settings")
def get_settings():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    out = {k: setting(db, k) for k in ["event_name", "event_subtitle", "event_date_text",
                                       "folio_start"]}
    out["ventas_cerradas"] = ventas_cerradas(db)
    out["seller_commission_pct"] = comision_general(db)
    out["door_code"] = setting(db, "door_code")
    out.update(flyer_info(db))
    return jsonify(out)

@app.post("/api/admin/reset")
def admin_reset():
    """Borra todo desde el panel. Existe porque resetear el propio evento no debería
    obligar a entrar a Railway y editar variables de servidor: eso ya falló dos veces
    en la práctica y deja al organizador atorado con datos de prueba a la vista.

    El candado no es un "¿seguro?" —a eso se le da que sí sin leer—: hay que escribir
    la palabra exacta. Y hay que ser el admin PRINCIPAL (el de la variable de
    entorno): un segundo admin invitado no puede borrarle el evento a nadie."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    principal = (os.environ.get("ADMIN_USER") or "admin").strip()
    if s["admin"]["username"] != principal:
        return jsonify(error=f"Solo {principal} puede borrar el sistema"), 403
    if str((request.json or {}).get("confirmar", "")).strip().upper() != "BORRAR TODO":
        return jsonify(error='Escribe exactamente: BORRAR TODO'), 400
    db = get_db()
    # el MISMO criterio del Resumen (sin invitados y sin anulados): si el panel dice
    # 55 boletos, el aviso de borrado no puede decir 62
    antes = db.execute(
        f"SELECT COUNT(*) c FROM tickets WHERE status!='void' AND {NOT_GUEST}").fetchone()["c"]
    vend = db.execute("SELECT COUNT(*) c FROM sellers WHERE hidden=0 AND deleted=0").fetchone()["c"]
    borrar_todo(db, principal, f"Sistema borrado desde el panel por {principal}")
    # el vendedor de invitados se recrea aquí mismo: si se esperara al próximo
    # arranque, los boletos de cortesía ya entregados dejarían de escanear
    guest = (os.environ.get("GUEST_SELLER_CODE") or "").strip()
    if guest and re.fullmatch(r"\d{4,6}", guest):
        db.execute("INSERT INTO sellers(name, code, active, hidden, created_at) "
                   "VALUES(?,?,1,1,?)", ("Invitados", guest, now_iso()))
        db.commit()
    return jsonify(ok=True, boletos=antes, vendedores=vend)

@app.post("/api/admin/door-code")
def door_code():
    """Genera o apaga la clave del escáner de la puerta.

    Generar una nueva (o apagarla) TAMBIÉN cierra las sesiones de escáner abiertas:
    rotar la clave debe sacar al staff viejo, o rotarla no protege nada."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    accion = str((request.json or {}).get("accion", "")).strip()
    if accion == "generar":
        clave = f"{secrets.randbelow(1000000):06d}"
        set_setting(db, "door_code", clave)
        db.execute("DELETE FROM sessions WHERE role='scanner'")
        audit(db, s["admin"]["username"], "ajustes",
              "Generó una clave nueva para el escáner de la puerta")
        db.commit()
        return jsonify(ok=True, door_code=clave)
    if accion == "apagar":
        set_setting(db, "door_code", "")
        db.execute("DELETE FROM sessions WHERE role='scanner'")
        audit(db, s["admin"]["username"], "ajustes",
              "Apagó la clave del escáner: solo los admins pueden escanear")
        db.commit()
        return jsonify(ok=True, door_code="")
    return jsonify(error="accion inválida"), 400

@app.post("/api/admin/ventas")
def toggle_ventas():
    """Apaga o enciende la boletera de los vendedores.

    Es para la noche de la fiesta: se cierra un par de horas antes para poder cortar
    cuentas sabiendo que el total ya no se mueve. Deliberadamente NO toca el escáner
    —la gente tiene que seguir entrando— ni el panel de admin. Queda en Movimientos
    porque es de las cosas más gordas que se pueden hacer aquí."""
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    cerrar = bool((request.json or {}).get("cerrar"))
    antes = ventas_cerradas(db)
    set_setting(db, "ventas_cerradas", "1" if cerrar else "0")
    if antes != cerrar:
        audit(db, s["admin"]["username"], "catalogo",
              "CERRÓ las ventas: los vendedores ya no pueden generar boletos"
              if cerrar else "Reabrió las ventas: los vendedores vuelven a generar")
    db.commit()
    return jsonify(ok=True, ventas_cerradas=cerrar)

def _clamp(v, lo, hi, default):
    try:
        return max(lo, min(hi, float(v)))
    except (TypeError, ValueError):
        return default

@app.post("/api/admin/settings")
def save_settings():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    db = get_db()
    b = request.json or {}
    changed = []
    for k in ["event_name", "event_subtitle", "event_date_text"]:
        if k in b:
            set_setting(db, k, str(b[k]).strip())
            changed.append(k)
    if "folio_start" in b:
        # número del primer folio: sirve para no empezar en 0001 y que no se
        # sepa cuántos boletos llevamos vendidos
        try:
            fs = max(1, min(9999, int(b["folio_start"])))
        except (TypeError, ValueError):
            fs = 1
        set_setting(db, "folio_start", str(fs))
        changed.append(f"folio inicial {fs:04d}")
    if "seller_commission_pct" in b:
        # el porcentaje que se aplica a quien no tenga uno propio
        try:
            pc = max(0.0, min(100.0, float(b["seller_commission_pct"])))
        except (TypeError, ValueError):
            return jsonify(error="La comisión debe ser un número entre 0 y 100"), 400
        set_setting(db, "seller_commission_pct", str(pc))
        changed.append(f"comisión general {pc:g}%")
    # posición/zoom de cada flyer (reposicionar sin volver a subir la imagen)
    for v in FLYER_VARIANTS:
        if f"flyer_focus_{v}" in b:
            set_setting(db, f"flyer_focus_{v}", _clamp(b[f"flyer_focus_{v}"], 0, 1, 0.5))
            changed.append(f"flyer_focus_{v}")
        if f"flyer_scale_{v}" in b:
            set_setting(db, f"flyer_scale_{v}", _clamp(b[f"flyer_scale_{v}"], 1, 3, 1))
            changed.append(f"flyer_scale_{v}")
    audit(db, s["admin"]["username"], "ajustes", f"Actualizó ajustes: {', '.join(changed)}")
    db.commit()
    return jsonify(ok=True)

@app.post("/api/admin/flyer")
def upload_flyer():
    s = require_admin()
    if not s:
        return jsonify(error="sin sesión"), 401
    f = request.files.get("flyer")
    if not f:
        return jsonify(error="Sube una imagen"), 400
    variant = request.form.get("variant")
    if variant not in FLYER_VARIANTS:
        return jsonify(error="Tipo de flyer inválido"), 400
    ext = os.path.splitext(f.filename or "")[1].lower()
    mimes = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".webp": "image/webp"}
    if ext not in mimes:
        return jsonify(error="Usa PNG, JPG o WEBP"), 400
    db = get_db()
    # El flyer se guarda EN LA BASE DE DATOS (base64), no en disco → no depende de volúmenes.
    raw = f.read()
    set_setting(db, f"flyer_data_{variant}", base64.b64encode(raw).decode())
    set_setting(db, f"flyer_mime_{variant}", mimes[ext])
    set_setting(db, f"flyer_focus_{variant}", _clamp(request.form.get("flyer_focus"), 0, 1, 0.5))
    set_setting(db, f"flyer_scale_{variant}", _clamp(request.form.get("flyer_scale"), 1, 3, 1))
    audit(db, s["admin"]["username"], "ajustes", f"Subió el flyer {FLYER_LABEL[variant]}")
    db.commit()
    return jsonify(ok=True)

@app.get("/flyer")
def serve_flyer():
    """Sirve el flyer del tipo pedido (?v=uady|externo|vip|grupo10|ultravip), con
    respaldo en cadena hasta el flyer legado de una sola imagen."""
    db = get_db()
    v = request.args.get("v")
    if v not in FLYER_VARIANTS:
        v = "externo"
    data = mime = None
    for vv in _flyer_chain(v):
        data = setting(db, f"flyer_data_{vv}")
        if data:
            mime = setting(db, f"flyer_mime_{vv}") or "image/png"
            break
    if not data:
        data = setting(db, "flyer_data")
        mime = setting(db, "flyer_mime") or "image/png"
    if not data:
        return "", 404
    resp = Response(base64.b64decode(data), mimetype=mime)
    resp.headers["Cache-Control"] = "no-cache"
    return resp

# ---------------------------------------------------------------- estáticos

def _version_estaticos():
    """Sello de esta versión del CSS y el JS: la fecha del archivo más nuevo.

    Sin esto, un arreglo subido a media venta NO llega a los teléfonos: Safari se
    queda con el CSS y el JS que descargó la primera vez y no vuelve a pedirlos,
    aunque el HTML sí se revalide. Con el sello pegado a la URL, cada versión es un
    archivo distinto para el navegador y no hay nada viejo que reusar."""
    ultimo = 0
    for carpeta in ("css", "js"):
        d = os.path.join(PUBLIC, carpeta)
        if not os.path.isdir(d):
            continue
        for f in os.listdir(d):
            try:
                ultimo = max(ultimo, int(os.path.getmtime(os.path.join(d, f))))
            except OSError:
                pass
    return str(ultimo)

ASSET_V = _version_estaticos()
_RE_ASSET = re.compile(r'(?P<a>(?:src|href)=")(?P<p>/(?:css|js)/[^"?]+)"')

def pagina(nombre):
    """Sirve una página pegándole el sello de versión al CSS y al JS que carga."""
    with open(os.path.join(PUBLIC, nombre), encoding="utf-8") as fh:
        html = fh.read()
    html = _RE_ASSET.sub(lambda m: f'{m.group("a")}{m.group("p")}?v={ASSET_V}"', html)
    return Response(html, mimetype="text/html")

@app.get("/")
def index():
    return pagina("index.html")

@app.get("/admin")
def admin_page():
    return pagina("admin.html")

@app.get("/scan")
def scan_page():
    return pagina("scan.html")

@app.get("/sw.js")
def service_worker():
    # service worker de autodestrucción: limpia el escáner viejo de celulares que lo instalaron
    resp = send_from_directory(PUBLIC, "sw.js", mimetype="application/javascript")
    resp.headers["Cache-Control"] = "no-cache"
    return resp

@app.get("/<path:path>")
def static_files(path):
    # las páginas también se piden por su nombre de archivo (/admin.html): que pasen
    # por el mismo sellado, o cargarían el CSS viejo
    if path in ("index.html", "admin.html", "scan.html"):
        return pagina(path)
    return send_from_directory(PUBLIC, path)

# ---------------------------------------------------------------- arranque

init_db()
sync_excel()
threading.Thread(target=backup_loop, daemon=True).start()
threading.Thread(target=_excel_worker, daemon=True).start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8756"))
    print(f"[OnFire] Vendedores:   http://localhost:{port}/")
    print(f"[OnFire] Admin:        http://localhost:{port}/admin")
    print(f"[OnFire] Excel en vivo: {XLSX}")
    app.run(host="0.0.0.0", port=port, threaded=True)
